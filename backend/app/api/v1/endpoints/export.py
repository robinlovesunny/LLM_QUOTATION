"""
导出服务API端点
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel
from typing import Optional, List
from io import BytesIO
import logging

from app.core.database import get_db
from app.crud.quote import QuoteCRUD
from app.services.excel_exporter import get_excel_exporter
from app.services.oss_uploader import get_oss_uploader

logger = logging.getLogger(__name__)
router = APIRouter()


# ========== Schemas ==========
class ExportRequest(BaseModel):
    """导出请求"""
    quote_id: str
    template_type: str = "standard"  # standard/competitor/simplified


class ExportResponse(BaseModel):
    """导出响应"""
    download_url: str
    message: str
    file_size: Optional[int] = None


class TemplateInfo(BaseModel):
    """模板信息"""
    id: str
    name: str
    description: str


# ========== API Endpoints ==========
@router.post("/excel", response_model=ExportResponse)
async def export_excel(
    request: ExportRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    导出Excel报价单
    
    Args:
        request: 导出请求
        db: 数据库会话
    
    Returns:
        下载链接和消息
    """
    try:
        # 1. 获取报价单数据
        quote = await QuoteCRUD.get_quote(db, request.quote_id)
        if not quote:
            raise HTTPException(status_code=404, detail="报价单不存在")
        
        # 2. 获取报价明细
        items = await QuoteCRUD.get_quote_items(db, request.quote_id)
        if not items:
            raise HTTPException(status_code=400, detail="报价单无明细数据")
        
        # 3. 生成Excel文件
        exporter = get_excel_exporter()
        
        if request.template_type == "standard":
            file_content = await exporter.generate_standard_quote(quote, items)
        elif request.template_type == "competitor":
            # 竞品对比版本(需要额外数据)
            file_content = await exporter.generate_competitor_comparison(
                quote, items, {}
            )
        elif request.template_type == "simplified":
            file_content = await exporter.generate_simplified_quote(quote, items)
        else:
            raise HTTPException(status_code=400, detail="不支持的模板类型")
        
        # 4. 上传到OSS
        uploader = get_oss_uploader()
        download_url = await uploader.upload_quote_file(
            file_content,
            request.quote_id,
            "xlsx"
        )
        
        if not download_url:
            raise HTTPException(status_code=500, detail="文件上传失败")
        
        logger.info(f"报价单导出成功: {request.quote_id}")
        
        return ExportResponse(
            download_url=download_url,
            message="Excel导出成功",
            file_size=len(file_content)
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出Excel失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.post("/pdf", response_model=ExportResponse)
async def export_pdf(
    request: ExportRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    导出PDF报价单
    
    Args:
        request: 导出请求
        db: 数据库会话
    
    Returns:
        下载链接和消息
    """
    # PDF导出可以后续通过Excel转换实现
    # 目前返回提示信息
    return ExportResponse(
        download_url="",
        message="PDF导出功能将在后续版本提供"
    )


@router.get("/templates")
async def get_templates():
    """
    获取可用模板列表
    
    Returns:
        模板列表
    """
    return [
        TemplateInfo(
            id="standard",
            name="标准报价单",
            description="包含完整产品信息、价格明细、折扣说明的标准格式报价单"
        ),
        TemplateInfo(
            id="competitor",
            name="竞品对比版",
            description="包含火山引擎竞品价格对比的报价单"
        ),
        TemplateInfo(
            id="simplified",
            name="简化版",
            description="简化版报价单,仅包含产品名称、数量和价格"
        )
    ]


@router.get("/download/{quote_id}")
async def download_excel_direct(
    quote_id: str,
    template_type: str = "standard",
    db: AsyncSession = Depends(get_db)
):
    """
    直接下载Excel报价单（不通过OSS）
    
    直接返回Excel文件流，适合小文件快速下载
    
    Args:
        quote_id: 报价单ID
        template_type: 模板类型 (standard/competitor/simplified)
        db: 数据库会话
    
    Returns:
        Excel文件流
    """
    try:
        # 1. 获取报价单数据
        quote = await QuoteCRUD.get_quote(db, quote_id)
        if not quote:
            raise HTTPException(status_code=404, detail="报价单不存在")
        
        # 2. 获取报价明细
        items = await QuoteCRUD.get_quote_items(db, quote_id)
        if not items:
            raise HTTPException(status_code=400, detail="报价单无明细数据")
        
        # 3. 生成Excel文件
        exporter = get_excel_exporter()
        
        if template_type == "standard":
            file_content = await exporter.generate_standard_quote(quote, items)
        elif template_type == "competitor":
            file_content = await exporter.generate_competitor_comparison(quote, items, {})
        elif template_type == "simplified":
            file_content = await exporter.generate_simplified_quote(quote, items)
        else:
            raise HTTPException(status_code=400, detail="不支持的模板类型")
        
        # 4. 返回文件流
        filename = f"报价单_{quote.quote_no}.xlsx"
        
        return StreamingResponse(
            BytesIO(file_content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
            }
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"下载Excel失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"下载失败: {str(e)}")


@router.get("/preview/{quote_id}")
async def preview_quote_data(
    quote_id: str,
    db: AsyncSession = Depends(get_db)
):
    """
    预览报价单数据（JSON格式）
    
    在导出前预览报价单内容，确认数据正确
    
    Args:
        quote_id: 报价单ID
        db: 数据库会话
    
    Returns:
        报价单预览数据
    """
    try:
        quote = await QuoteCRUD.get_quote(db, quote_id)
        if not quote:
            raise HTTPException(status_code=404, detail="报价单不存在")
        
        items = await QuoteCRUD.get_quote_items(db, quote_id)
        
        return {
            "quote": {
                "quote_id": str(quote.quote_id),
                "quote_no": quote.quote_no,
                "customer_name": quote.customer_name,
                "project_name": quote.project_name,
                "status": quote.status,
                "total_amount": float(quote.total_amount) if quote.total_amount else 0,
                "currency": quote.currency,
                "created_at": quote.created_at.isoformat() if quote.created_at else None,
                "valid_until": quote.valid_until.isoformat() if quote.valid_until else None
            },
            "items": [
                {
                    "product_name": item.product_name,
                    "product_code": item.product_code,
                    "region": item.region,
                    "quantity": item.quantity,
                    "duration_months": item.duration_months,
                    "original_price": float(item.original_price) if item.original_price else 0,
                    "discount_rate": float(item.discount_rate) if item.discount_rate else 1,
                    "final_price": float(item.final_price) if item.final_price else 0
                }
                for item in items
            ],
            "summary": {
                "item_count": len(items),
                "total_original": sum(float(item.original_price or 0) for item in items),
                "total_final": sum(float(item.final_price or 0) for item in items)
            }
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"预览报价单失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"预览失败: {str(e)}")


class BatchExportRequest(BaseModel):
    """批量导出请求"""
    quote_ids: List[str]
    template_type: str = "standard"


class BatchExportResult(BaseModel):
    """批量导出结果"""
    success_count: int
    failed_count: int
    results: List[dict]


class QuotePreviewRequest(BaseModel):
    """报价预览导出请求"""
    customerInfo: dict  # 客户信息
    selectedModels: List[dict]  # 已选模型
    modelConfigs: dict  # 模型配置
    specDiscounts: Optional[dict] = {}  # 规格折扣
    dailyUsages: Optional[dict] = {}  # 日估计用量
    priceUnit: Optional[str] = 'thousand'  # 价格单位: 'thousand'(千Token) 或 'million'(百万Token)


@router.post("/batch", response_model=BatchExportResult)
async def batch_export(
    request: BatchExportRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    批量导出报价单
    
    一次性导出多个报价单并上传到OSS
    
    Args:
        request: 批量导出请求
        db: 数据库会话
    
    Returns:
        批量导出结果
    """
    try:
        results = []
        success_count = 0
        failed_count = 0
        
        exporter = get_excel_exporter()
        uploader = get_oss_uploader()
        
        for quote_id in request.quote_ids:
            try:
                quote = await QuoteCRUD.get_quote(db, quote_id)
                if not quote:
                    results.append({
                        "quote_id": quote_id,
                        "success": False,
                        "error": "报价单不存在"
                    })
                    failed_count += 1
                    continue
                
                items = await QuoteCRUD.get_quote_items(db, quote_id)
                if not items:
                    results.append({
                        "quote_id": quote_id,
                        "success": False,
                        "error": "报价单无明细数据"
                    })
                    failed_count += 1
                    continue
                
                # 生成并上传
                excel_bytes, oss_url = await exporter.generate_and_upload(
                    quote, items, request.template_type
                )
                
                results.append({
                    "quote_id": quote_id,
                    "quote_no": quote.quote_no,
                    "success": True,
                    "download_url": oss_url,
                    "file_size": len(excel_bytes)
                })
                success_count += 1
                
            except Exception as e:
                results.append({
                    "quote_id": quote_id,
                    "success": False,
                    "error": str(e)
                })
                failed_count += 1
        
        return BatchExportResult(
            success_count=success_count,
            failed_count=failed_count,
            results=results
        )
    
    except Exception as e:
        logger.error(f"批量导出失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"批量导出失败: {str(e)}")


@router.post("/preview")
async def export_quote_preview(
    request: QuotePreviewRequest
):
    """
    导出报价预览Excel
    
    接收前端传来的报价数据，生成Excel文件并返回下载链接
    使用category-specific rendering logic for proper table structure
    """
    import os
    import uuid
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from app.services.excel_exporter import (
        group_models_by_category,
        render_category_section,
        CATEGORY_CONFIG
    )
    
    try:
        # Step 1: Extract request data
        customer_info = request.customerInfo
        selected_models = request.selectedModels
        model_configs = request.modelConfigs
        spec_discounts = request.specDiscounts or {}
        daily_usages = request.dailyUsages or {}
        price_unit = request.priceUnit or 'thousand'
        
        # Extract customer info fields
        customer_name = customer_info.get('customerName', '')
        quote_date = customer_info.get('quoteDate', '')
        valid_until = customer_info.get('validUntil', '')
        discount_percent = customer_info.get('discountPercent', 0)
        
        # Debug logging
        logger.info(f"导出请求数据 - selectedModels: {len(selected_models)} models")
        logger.info(f"导出请求数据 - modelConfigs keys: {list(model_configs.keys())}")
        logger.info(f"导出请求数据 - priceUnit: {price_unit}")
        
        # DEBUG: Log sample model and config data
        if selected_models:
            import json
            logger.info(f"=== EXPORT DEBUG START ===")
            logger.info(f"DEBUG - First selected_model: {json.dumps(selected_models[0], ensure_ascii=False, default=str)}")
            first_key = list(model_configs.keys())[0] if model_configs else None
            if first_key:
                config_sample = model_configs[first_key]
                logger.info(f"DEBUG - First model_config (key={first_key}): {json.dumps(config_sample, ensure_ascii=False, default=str)}")
            logger.info(f"=== EXPORT DEBUG END ===")
        
        # Step 2: Group models by category
        grouped_models = group_models_by_category(selected_models, model_configs)
        logger.info(f"Grouped models into {len(grouped_models)} categories: {list(grouped_models.keys())}")
        
        # Step 3: Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "报价清单"
        
        # Step 4: Render title and customer info
        # Title styling
        title_font = Font(name='微软雅黑', size=16, bold=True)
        
        # Write title
        ws.merge_cells('A1:H1')
        ws['A1'] = '阿里云大模型产品报价清单'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Write customer info
        ws['A3'] = '客户名称：'
        ws['B3'] = customer_name
        ws['D3'] = '报价日期：'
        ws['E3'] = quote_date
        ws['G3'] = '有效期：'
        ws['H3'] = valid_until
        
        # Step 5: Initialize current_row
        current_row = 5
        
        # Step 6: For each category in CATEGORY_CONFIG order, render category section
        # Define category order (same as CATEGORY_CONFIG order field)
        category_order = sorted(
            CATEGORY_CONFIG.keys(),
            key=lambda k: CATEGORY_CONFIG[k]['order']
        )
        
        for category_key in category_order:
            # Check if category has models in grouped_models
            if category_key not in grouped_models:
                continue
            
            category_data = grouped_models[category_key]
            models = category_data.get('models', [])
            
            if not models:
                continue
            
            # Call render_category_section()
            logger.info(f"Rendering category: {category_key} with {len(models)} models")
            current_row = render_category_section(
                ws,
                category_key,
                models,
                current_row,
                price_unit,
                spec_discounts,
                daily_usages,
                discount_percent
            )
        
        # Step 7: Render quote notes
        current_row += 1  # Add extra spacing before notes
        
        ws.cell(row=current_row, column=1, value='报价说明：').font = Font(name='微软雅黑', bold=True)
        current_row += 1
        
        ws.cell(row=current_row, column=1, value='• 以上价格均为人民币（CNY）计价')
        current_row += 1
        
        ws.cell(row=current_row, column=1, value='• Token计费模型按实际调用量结算')
        current_row += 1
        
        if discount_percent > 0:
            discount_display = (100 - discount_percent) / 10
            ws.cell(row=current_row, column=1, value=f'• 本报价单默认折扣: {discount_display:.1f}折')
        
        # Step 8: Save file and return response
        exports_dir = "exports"
        os.makedirs(exports_dir, exist_ok=True)
        
        filename = f"报价单_{customer_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}.xlsx"
        filepath = os.path.join(exports_dir, filename)
        wb.save(filepath)
        
        logger.info(f"Excel file saved successfully: {filename}")
        
        return {
            "success": True,
            "filename": filename,
            "message": "报价单生成成功"
        }
    
    except Exception as e:
        logger.error(f"导出报价预览失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/download/file/{filename}")
async def download_export_file(filename: str):
    """
    下载导出的文件
    """
    import os
    from fastapi.responses import FileResponse
    
    filepath = os.path.join("exports", filename)
    
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="文件不存在")
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

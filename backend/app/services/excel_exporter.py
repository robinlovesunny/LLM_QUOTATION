"""
Excel导出服务 - 生成Excel格式报价单
"""
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime
from decimal import Decimal
from pathlib import Path
import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.models.quote import QuoteSheet, QuoteItem
from app.services.oss_uploader import get_oss_uploader

logger = logging.getLogger(__name__)


# Category configuration with icons, names, and pricing types
CATEGORY_CONFIG = {
    'text_qwen': {
        'name': '文本生成-通义千问',
        'icon': '💬',
        'price_type': 'token',
        'order': 1
    },
    'text_qwen_opensource': {
        'name': '文本生成-通义千问-开源版',
        'icon': '📝',
        'price_type': 'token',
        'order': 2
    },
    'text_thirdparty': {
        'name': '文本生成-第三方模型',
        'icon': '🤖',
        'price_type': 'token',
        'order': 3
    },
    'image_gen': {
        'name': '图像生成',
        'icon': '🎨',
        'price_type': 'image',
        'order': 4
    },
    'image_gen_thirdparty': {
        'name': '图像生成-第三方模型',
        'icon': '🖼️',
        'price_type': 'image',
        'order': 5
    },
    'tts': {
        'name': '语音合成',
        'icon': '🔊',
        'price_type': 'character',
        'order': 6
    },
    'asr': {
        'name': '语音识别与翻译',
        'icon': '🎤',
        'price_type': 'audio',
        'order': 7
    },
    'video_gen': {
        'name': '视频生成',
        'icon': '🎬',
        'price_type': 'video',
        'order': 8
    },
    'text_embedding': {
        'name': '文本向量',
        'icon': '📊',
        'price_type': 'token',
        'order': 9
    },
    'multimodal_embedding': {
        'name': '多模态向量',
        'icon': '🌐',
        'price_type': 'token',
        'order': 10
    },
    'text_nlu': {
        'name': '文本分类抽取排序',
        'icon': '🔍',
        'price_type': 'token',
        'order': 11
    },
    'industry': {
        'name': '行业模型',
        'icon': '🏭',
        'price_type': 'token',
        'order': 12
    }
}

# Unit mapping for non-token pricing dimension codes
UNIT_MAP = {
    'character': '字符',
    'audio_second': '秒',
    'video_second': '秒',
    'image_count': '张'
}


def get_category_config(category_key: str) -> Optional[Dict[str, Any]]:
    """
    Get category configuration by key.
    
    Args:
        category_key: Category identifier (e.g., 'text_qwen', 'image_gen')
    
    Returns:
        Category configuration dict or None if not found
    """
    return CATEGORY_CONFIG.get(category_key)


def classify_model(model: Dict[str, Any]) -> str:
    """
    Classify a model into one of the 12 categories.
    
    Args:
        model: Model data with category, sub_category, model_code, name, etc.
    
    Returns:
        Category key (e.g., 'text_qwen', 'image_gen')
    
    Logic:
        1. Check if model.category or model.sub_category matches a category key
        2. If not, classify by model name patterns:
           - wanx/flux/stable-diffusion/qwen-image/image-edit → image_gen
           - t2v/i2v/wan2* → video_gen
           - *-tts/cosyvoice → tts
           - *-asr/paraformer/sensevoice → asr
           - *embedding* → text_embedding
        3. Default to text_qwen if no match
    """
    # Step 1: Check category and sub_category fields
    category = (model.get('category') or '').lower()
    sub_category = (model.get('sub_category') or '').lower()
    
    # Check if category or sub_category matches a known category key
    if category in CATEGORY_CONFIG:
        return category
    if sub_category in CATEGORY_CONFIG:
        return sub_category
    
    # Step 2: Apply name pattern matching
    # Get model name from various possible fields
    model_name = (
        model.get('model_code') or 
        model.get('name') or 
        model.get('model_name') or
        ''
    ).lower()
    
    # Image generation patterns
    if any(pattern in model_name for pattern in ['wanx', 'flux', 'stable-diffusion', 'qwen-image', 'image-edit']):
        return 'image_gen'
    
    # Video generation patterns
    if any(pattern in model_name for pattern in ['t2v', 'i2v']) or model_name.startswith('wan2'):
        return 'video_gen'
    
    # TTS (Text-to-Speech) patterns
    if model_name.endswith('-tts') or 'cosyvoice' in model_name:
        return 'tts'
    
    # ASR (Automatic Speech Recognition) patterns
    if model_name.endswith('-asr') or 'paraformer' in model_name or 'sensevoice' in model_name:
        return 'asr'
    
    # Embedding patterns
    if 'embedding' in model_name:
        return 'text_embedding'
    
    # Step 3: Default to text_qwen
    return 'text_qwen'


def extract_prices(spec: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract prices from spec data, supporting both new and legacy formats.
    
    Args:
        spec: Spec data with prices array or legacy price fields
    
    Returns:
        {
            'input_price': float | None,
            'output_price': float | None,
            'non_token_price': float | None,
            'price_unit': str | None,
            'dimension_code': str | None
        }
    
    Logic:
        1. If spec has 'prices' array:
           - Extract input from dimension_code: input, input_token, input_token_image
           - Extract output from dimension_code: output, output_token, output_token_thinking
           - Extract non-token from dimension_code: character, audio_second, video_second, image_count
           - Map dimension_code to unit using UNIT_MAP
        2. Else fall back to legacy fields: input_price, output_price
    """
    result = {
        'input_price': None,
        'output_price': None,
        'non_token_price': None,
        'price_unit': None,
        'dimension_code': None
    }
    
    # Check if spec has prices array (new format)
    prices = spec.get('prices')
    if prices and isinstance(prices, list):
        # Process each price entry in the array
        for price_entry in prices:
            if not isinstance(price_entry, dict):
                continue
            
            dimension_code = price_entry.get('dimension_code', '').lower()
            unit_price = price_entry.get('unit_price')
            
            # Skip if no valid price
            if unit_price is None:
                continue
            
            # Convert to float if needed
            try:
                unit_price = float(unit_price)
            except (ValueError, TypeError):
                continue
            
            # Extract input prices
            if dimension_code in ['input', 'input_token', 'input_token_image']:
                result['input_price'] = unit_price
            
            # Extract output prices
            elif dimension_code in ['output', 'output_token', 'output_token_thinking']:
                result['output_price'] = unit_price
            
            # Extract non-token prices
            elif dimension_code in ['character', 'audio_second', 'video_second', 'image_count']:
                result['non_token_price'] = unit_price
                result['dimension_code'] = dimension_code
                # Map dimension_code to unit label
                result['price_unit'] = UNIT_MAP.get(dimension_code)
    
    # Fall back to legacy fields if no prices were extracted
    if result['input_price'] is None and result['output_price'] is None and result['non_token_price'] is None:
        # Try legacy input_price field
        legacy_input = spec.get('input_price')
        if legacy_input is not None:
            try:
                result['input_price'] = float(legacy_input)
            except (ValueError, TypeError):
                pass
        
        # Try legacy output_price field
        legacy_output = spec.get('output_price')
        if legacy_output is not None:
            try:
                result['output_price'] = float(legacy_output)
            except (ValueError, TypeError):
                pass
    
    return result


def convert_price_unit(price: Optional[float], price_unit: str) -> Tuple[Optional[float], str]:
    """
    Convert token price based on unit preference.
    
    Args:
        price: Original price (per 千Token), can be None
        price_unit: 'thousand' or 'million'
    
    Returns:
        Tuple of (converted_price, unit_label)
        - If price is None, returns (None, unit_label)
        - If price_unit == 'thousand': returns (price, '千Token')
        - If price_unit == 'million': returns (price * 1000, '百万Token')
    
    Examples:
        >>> convert_price_unit(0.5, 'thousand')
        (0.5, '千Token')
        >>> convert_price_unit(0.5, 'million')
        (500.0, '百万Token')
        >>> convert_price_unit(None, 'thousand')
        (None, '千Token')
    """
    # Handle None prices gracefully
    if price is None:
        unit_label = '千Token' if price_unit == 'thousand' else '百万Token'
        return (None, unit_label)
    
    # Convert based on price_unit
    if price_unit == 'million':
        return (price * 1000, '百万Token')
    else:
        # Default to 'thousand' for any other value
        return (price, '千Token')


def get_spec_discount(
    model_id: str,
    spec_id: str,
    spec_discounts: Dict[str, Dict[str, float]],
    global_discount: float
) -> float:
    """
    Get the discount for a specific spec, with spec-level taking precedence over global.
    
    Args:
        model_id: Model identifier (can be model_code or id)
        spec_id: Spec identifier
        spec_discounts: Nested dict of spec-level discounts {model_id: {spec_id: discount_percent}}
        global_discount: Global discount percentage (0-100)
    
    Returns:
        Discount percentage (0-100) - spec-level if exists, otherwise global_discount
    
    Examples:
        >>> get_spec_discount('model1', 'spec1', {'model1': {'spec1': 10.0}}, 5.0)
        10.0
        >>> get_spec_discount('model1', 'spec2', {'model1': {'spec1': 10.0}}, 5.0)
        5.0
        >>> get_spec_discount('model2', 'spec1', {}, 5.0)
        5.0
    """
    # Check if spec_discounts has this model_id
    if model_id in spec_discounts:
        model_discounts = spec_discounts[model_id]
        # Check if this spec_id has a custom discount
        if isinstance(model_discounts, dict) and spec_id in model_discounts:
            return model_discounts[spec_id]
    
    # Fall back to global discount
    return global_discount


def has_any_discount(
    spec_discounts: Dict[str, Dict[str, float]],
    global_discount: float
) -> bool:
    """
    Check if any discounts are present (global or spec-level).
    
    Args:
        spec_discounts: Nested dict of spec-level discounts {model_id: {spec_id: discount_percent}}
        global_discount: Global discount percentage (0-100)
    
    Returns:
        True if global_discount > 0 or any spec discount > 0, False otherwise
    
    Examples:
        >>> has_any_discount({}, 0.0)
        False
        >>> has_any_discount({}, 5.0)
        True
        >>> has_any_discount({'model1': {'spec1': 10.0}}, 0.0)
        True
        >>> has_any_discount({'model1': {'spec1': 0.0}}, 0.0)
        False
    """
    # Check if global discount is greater than 0
    if global_discount > 0:
        return True
    
    # Check if any spec-level discount is greater than 0
    if spec_discounts:
        for model_id, model_discounts in spec_discounts.items():
            if isinstance(model_discounts, dict):
                for spec_id, discount in model_discounts.items():
                    if discount > 0:
                        return True
    
    return False


def calculate_monthly_usage(daily_usage: str) -> str:
    """
    Calculate monthly usage from daily usage.
    
    Args:
        daily_usage: Daily usage value as string (e.g., "1000", "500.5")
    
    Returns:
        Formatted string with monthly usage (daily × 30)
        Returns "-" for empty/invalid input
    
    Examples:
        >>> calculate_monthly_usage("1000")
        "30000"
        >>> calculate_monthly_usage("500.5")
        "15015.0"
        >>> calculate_monthly_usage("")
        "-"
        >>> calculate_monthly_usage("invalid")
        "-"
        >>> calculate_monthly_usage("0")
        "0"
    """
    # Handle empty or None input
    if not daily_usage or daily_usage is None:
        return "-"
    
    # Handle string input that's just whitespace
    if isinstance(daily_usage, str) and not daily_usage.strip():
        return "-"
    
    try:
        # Parse to float
        daily_value = float(daily_usage)
        
        # Calculate monthly usage (daily × 30)
        monthly_value = daily_value * 30
        
        # Special case: zero should always be formatted as "0"
        if monthly_value == 0:
            return "0"
        
        # Format the result
        # Check if the original input had a decimal point to preserve formatting intent
        # Also check for 'e' or 'E' for scientific notation which implies decimal
        input_str = str(daily_usage).strip()
        has_decimal_in_input = '.' in input_str or 'e' in input_str.lower()
        
        # If the result is a whole number AND input didn't have decimal, return without decimal point
        if monthly_value == int(monthly_value) and not has_decimal_in_input:
            return str(int(monthly_value))
        else:
            return str(monthly_value)
    
    except (ValueError, TypeError):
        # Return "-" for invalid input that can't be converted to float
        return "-"


def calculate_monthly_cost(
    spec: Dict[str, Any],
    daily_usage: str,
    discount: float,
    price_type: str
) -> str:
    """
    Calculate monthly cost based on daily usage, discount, and price type.
    
    Args:
        spec: Spec data containing price information (from extract_prices result)
        daily_usage: Daily usage value as string (e.g., "1000", "500.5")
        discount: Discount percentage (0-100)
        price_type: Price type ('token' for token-based, or 'image'/'character'/'audio'/'video' for non-token)
    
    Returns:
        Formatted currency string (e.g., "¥1,234.56")
        Returns "-" for missing data or invalid input
    
    Logic:
        - Parse daily_usage to float
        - Calculate discount_rate = (100 - discount) / 100
        - For token models: cost = (input_price + output_price) × daily × 30 × discount_rate
        - For non-token models: cost = non_token_price × daily × 30 × discount_rate
        - Return formatted currency string
        - Handle missing data by returning "-"
    
    Examples:
        >>> spec = {'input_price': 0.04, 'output_price': 0.12}
        >>> calculate_monthly_cost(spec, "1000", 10.0, 'token')
        "¥4,320.00"
        >>> spec = {'non_token_price': 0.08}
        >>> calculate_monthly_cost(spec, "500", 0.0, 'image')
        "¥1,200.00"
        >>> calculate_monthly_cost({}, "1000", 0.0, 'token')
        "-"
        >>> calculate_monthly_cost(spec, "", 0.0, 'token')
        "-"
    
    Requirements: 6.3, 6.4, 6.5
    """
    # Handle empty or None daily_usage
    if not daily_usage or daily_usage is None:
        return "-"
    
    # Handle string input that's just whitespace
    if isinstance(daily_usage, str) and not daily_usage.strip():
        return "-"
    
    try:
        # Parse daily_usage to float
        daily_value = float(daily_usage)
        
        # Handle zero or negative daily usage
        if daily_value <= 0:
            return "-"
        
        # Calculate discount_rate = (100 - discount) / 100
        discount_rate = (100 - discount) / 100
        
        # Determine cost based on price_type
        if price_type == 'token':
            # For token models: cost = (input_price + output_price) × daily × 30 × discount_rate
            input_price = spec.get('input_price')
            output_price = spec.get('output_price')
            
            # Check if we have valid prices
            if input_price is None and output_price is None:
                return "-"
            
            # Use 0 for missing prices (some models may only have input or output)
            input_price = input_price if input_price is not None else 0.0
            output_price = output_price if output_price is not None else 0.0
            
            # Calculate total cost
            cost = (input_price + output_price) * daily_value * 30 * discount_rate
        
        else:
            # For non-token models: cost = non_token_price × daily × 30 × discount_rate
            non_token_price = spec.get('non_token_price')
            
            # Check if we have a valid price
            if non_token_price is None:
                return "-"
            
            # Calculate total cost
            cost = non_token_price * daily_value * 30 * discount_rate
        
        # Format as currency string with 2 decimal places and thousands separator
        # Using Chinese Yuan symbol
        return f"¥{cost:,.2f}"
    
    except (ValueError, TypeError) as e:
        # Return "-" for invalid input that can't be converted to float
        logger.debug(f"Error calculating monthly cost: {e}")
        return "-"


def render_category_header(
    ws,
    category_key: str,
    item_count: int,
    current_row: int,
    num_columns: int
) -> int:
    """
    Render a category header row in the Excel worksheet.
    
    Args:
        ws: Excel worksheet (openpyxl Worksheet object)
        category_key: Category identifier (e.g., 'text_qwen', 'image_gen')
        item_count: Number of items in this category
        current_row: Current row number to render the header
        num_columns: Total number of columns to merge across
    
    Returns:
        Next available row number (current_row + 1)
    
    Logic:
        1. Get category config (name, icon) using get_category_config()
        2. Merge cells across all columns (A to num_columns)
        3. Set cell value to "{icon} {name} (共{count}项)"
        4. Apply styling:
           - Font size: 12pt
           - Bold: True
           - Background color: #E7E6E6
           - Alignment: Left horizontal, Center vertical
        5. Return next row number
    
    Requirements: 2.1, 2.2, 2.3
    
    Examples:
        >>> # Render header for text_qwen category with 5 items at row 10, spanning 11 columns
        >>> next_row = render_category_header(ws, 'text_qwen', 5, 10, 11)
        >>> # Cell A10 will contain: "💬 文本生成-通义千问 (共5项)"
        >>> # Cells A10:K10 will be merged
        >>> # next_row will be 11
    """
    # Get category configuration
    config = get_category_config(category_key)
    
    # Handle invalid category key gracefully
    if config is None:
        logger.warning(f"Invalid category key: {category_key}, using default")
        config = {
            'name': category_key,
            'icon': '📋',
        }
    
    # Get category name and icon
    category_name = config['name']
    category_icon = config['icon']
    
    # Create the header text: "{icon} {name} (共{count}项)"
    header_text = f"{category_icon} {category_name} (共{item_count}项)"
    
    # Merge cells across all columns
    # Convert num_columns to column letter
    end_column_letter = get_column_letter(num_columns)
    merge_range = f"A{current_row}:{end_column_letter}{current_row}"
    ws.merge_cells(merge_range)
    
    # Get the merged cell (top-left cell)
    header_cell = ws[f"A{current_row}"]
    
    # Set cell value
    header_cell.value = header_text
    
    # Apply styling
    # Font: 12pt, bold, default color (black)
    header_cell.font = Font(name='微软雅黑', size=12, bold=True)
    
    # Background: #E7E6E6
    header_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Alignment: left horizontal, center vertical
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Return next row number
    return current_row + 1


def render_token_based_table(
    ws,
    models: List[Dict[str, Any]],
    start_row: int,
    price_unit: str,
    spec_discounts: Dict[str, Dict[str, float]],
    daily_usages: Dict[str, Dict[str, str]],
    global_discount: float
) -> int:
    """
    Render table for token-based models.
    
    Args:
        ws: Excel worksheet (openpyxl Worksheet object)
        models: List of model data with specs to render
        start_row: Starting row number for the table
        price_unit: 'thousand' or 'million' for token price unit
        spec_discounts: Nested dict of spec-level discounts {model_id: {spec_id: discount_percent}}
        daily_usages: Nested dict of daily usage values {model_id: {spec_id: usage_value}}
        global_discount: Global discount percentage (0-100)
    
    Returns:
        Next available row number after the table
    
    Columns (without discounts):
        序号, 模型名称, 模式, Token范围, 输入单价, 输出单价, 日估计用量, 预估月用量, 预估月费, 备注
    
    Columns (with discounts):
        序号, 模型名称, 模式, Token范围, 输入单价, 输出单价, 折扣, 折后输入, 折后输出, 日估计用量, 预估月用量, 预估月费, 备注
    
    Logic:
        1. Determine if any discounts exist (global or spec-level)
        2. Render appropriate headers
        3. For each model and spec:
           - Extract prices using extract_prices()
           - Convert prices using convert_price_unit()
           - Get spec discount (custom or global)
           - Calculate discounted prices
           - Get daily usage
           - Calculate monthly usage (daily × 30)
           - Calculate monthly cost ((input + output) × daily × 30 × discount_rate)
           - Render data row with appropriate formatting
    
    Requirements: 1.1, 5.3, 9.2, 9.3
    """
    current_row = start_row
    
    # Step 1: Determine if any discounts exist
    has_discount = has_any_discount(spec_discounts, global_discount)
    
    # Step 2: Render table headers
    if has_discount:
        # Headers with discount columns
        headers = [
            '序号', '模型名称', '模式', 'Token范围', 
            '输入单价', '输出单价', '折扣', '折后输入', '折后输出',
            '日估计用量', '预估月用量', '预估月费', '备注'
        ]
    else:
        # Headers without discount columns
        headers = [
            '序号', '模型名称', '模式', 'Token范围',
            '输入单价', '输出单价', 
            '日估计用量', '预估月用量', '预估月费', '备注'
        ]
    
    # Apply header styling: bold, white text, blue background #4472C4, center alignment, borders
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    current_row += 1
    
    # Step 3: Render data rows for each model and spec
    row_number = 1
    
    for model_data in models:
        model = model_data.get('model', {})
        specs = model_data.get('specs', [])
        
        # Get model identifier (try multiple field names)
        model_id = str(model.get('id') or model.get('model_code') or model.get('name') or '')
        model_name = model.get('model_name') or model.get('name') or model.get('model_code') or '-'
        
        for spec in specs:
            # Get spec identifier
            spec_id = str(spec.get('id') or spec.get('spec_id') or '')
            
            # Extract prices using extract_prices()
            price_data = extract_prices(spec)
            input_price = price_data['input_price']
            output_price = price_data['output_price']
            
            # Convert prices using convert_price_unit()
            converted_input, unit_label = convert_price_unit(input_price, price_unit)
            converted_output, _ = convert_price_unit(output_price, price_unit)
            
            # Get spec discount using get_spec_discount()
            discount = get_spec_discount(model_id, spec_id, spec_discounts, global_discount)
            
            # Calculate discounted prices if applicable
            discount_rate = (100 - discount) / 100
            discounted_input = converted_input * discount_rate if converted_input is not None else None
            discounted_output = converted_output * discount_rate if converted_output is not None else None
            
            # Get daily usage from daily_usages dict
            daily_usage = ''
            if model_id in daily_usages and spec_id in daily_usages[model_id]:
                daily_usage = daily_usages[model_id][spec_id]
            
            # Calculate monthly usage
            monthly_usage = calculate_monthly_usage(daily_usage)
            
            # Calculate monthly cost
            monthly_cost = calculate_monthly_cost(
                price_data,
                daily_usage,
                discount,
                'token'
            )
            
            # Get other spec fields
            mode = spec.get('mode') or '-'
            token_range = spec.get('token_tier') or spec.get('token_range') or '-'
            remark = spec.get('remark') or ''
            
            # Render data row
            col_idx = 1
            
            # 序号
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = row_number
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # 模型名称
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = model_name
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # 模式
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = mode
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # Token范围
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = token_range
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # 输入单价
            cell = ws.cell(row=current_row, column=col_idx)
            if converted_input is not None:
                cell.value = f"¥{converted_input:.4f}/{unit_label}"
            else:
                cell.value = '-'
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # 输出单价
            cell = ws.cell(row=current_row, column=col_idx)
            if converted_output is not None:
                cell.value = f"¥{converted_output:.4f}/{unit_label}"
            else:
                cell.value = '-'
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # Discount columns (if applicable)
            if has_discount:
                # 折扣
                cell = ws.cell(row=current_row, column=col_idx)
                if discount > 0:
                    cell.value = f"{discount:.1f}%"
                else:
                    cell.value = '-'
                cell.font = Font(name='微软雅黑', size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                col_idx += 1
                
                # 折后输入
                cell = ws.cell(row=current_row, column=col_idx)
                if discounted_input is not None:
                    cell.value = f"¥{discounted_input:.4f}/{unit_label}"
                else:
                    cell.value = '-'
                cell.font = Font(name='微软雅黑', size=10)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                col_idx += 1
                
                # 折后输出
                cell = ws.cell(row=current_row, column=col_idx)
                if discounted_output is not None:
                    cell.value = f"¥{discounted_output:.4f}/{unit_label}"
                else:
                    cell.value = '-'
                cell.font = Font(name='微软雅黑', size=10)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                col_idx += 1
            
            # 日估计用量
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = daily_usage if daily_usage else '-'
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # 预估月用量
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = monthly_usage
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # 预估月费
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = monthly_cost
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # 备注
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = remark
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            current_row += 1
            row_number += 1
    
    # Return next row number
    return current_row


def render_category_section(
    ws,
    category_key: str,
    models: List[Dict[str, Any]],
    current_row: int,
    price_unit: str,
    spec_discounts: Dict[str, Dict[str, float]],
    daily_usages: Dict[str, Dict[str, str]],
    global_discount: float
) -> int:
    """
    Render a complete category section with header and table.
    
    Args:
        ws: Excel worksheet (openpyxl Worksheet object)
        category_key: Category identifier (e.g., 'text_qwen', 'image_gen')
        models: List of model data with specs to render in this category
        current_row: Starting row number for the section
        price_unit: 'thousand' or 'million' for token price unit
        spec_discounts: Nested dict of spec-level discounts {model_id: {spec_id: discount_percent}}
        daily_usages: Nested dict of daily usage values {model_id: {spec_id: usage_value}}
        global_discount: Global discount percentage (0-100)
    
    Returns:
        Next available row number after the section (including spacing row)
    
    Logic:
        1. Get category config to determine price_type
        2. Calculate item count (total number of specs across all models)
        3. Determine number of columns based on price_type and discount presence
        4. Render category header using render_category_header()
        5. Determine if category is token-based or non-token from config.price_type
        6. Call appropriate table renderer:
           - render_token_based_table() if price_type == 'token'
           - render_non_token_table() otherwise
        7. Add spacing row after section
        8. Return next row number
    
    Requirements: 1.3
    
    Examples:
        >>> # Render a token-based category section
        >>> next_row = render_category_section(
        ...     ws, 'text_qwen', models, 10, 'thousand', {}, {}, 0.0
        ... )
        >>> # Will render: category header at row 10, table starting at row 11, spacing at end
        
        >>> # Render a non-token category section
        >>> next_row = render_category_section(
        ...     ws, 'image_gen', models, 20, 'thousand', {}, {}, 5.0
        ... )
        >>> # Will render: category header at row 20, table starting at row 21, spacing at end
    """
    # Step 1: Get category configuration
    config = get_category_config(category_key)
    
    # Handle invalid category key gracefully
    if config is None:
        logger.warning(f"Invalid category key: {category_key}, using default token type")
        config = {
            'name': category_key,
            'icon': '📋',
            'price_type': 'token'
        }
    
    # Step 2: Calculate item count (total number of specs across all models)
    item_count = 0
    for model_data in models:
        specs = model_data.get('specs', [])
        item_count += len(specs)
    
    # Step 3: Determine number of columns based on price_type and discount presence
    has_discount = has_any_discount(spec_discounts, global_discount)
    price_type = config.get('price_type', 'token')
    
    if price_type == 'token':
        # Token-based table columns
        if has_discount:
            # 序号, 模型名称, 模式, Token范围, 输入单价, 输出单价, 折扣, 折后输入, 折后输出, 日估计用量, 预估月用量, 预估月费, 备注
            num_columns = 13
        else:
            # 序号, 模型名称, 模式, Token范围, 输入单价, 输出单价, 日估计用量, 预估月用量, 预估月费, 备注
            num_columns = 10
    else:
        # Non-token table columns
        if has_discount:
            # 序号, 模型名称, 单价, 单位, 折扣, 折后单价, 日估计用量, 预估月用量, 预估月费
            num_columns = 9
        else:
            # 序号, 模型名称, 单价, 单位, 日估计用量, 预估月用量, 预估月费
            num_columns = 7
    
    # Step 4: Render category header
    current_row = render_category_header(
        ws,
        category_key,
        item_count,
        current_row,
        num_columns
    )
    
    # Step 5 & 6: Determine table type and call appropriate renderer
    if price_type == 'token':
        # Render token-based table
        current_row = render_token_based_table(
            ws,
            models,
            current_row,
            price_unit,
            spec_discounts,
            daily_usages,
            global_discount
        )
    else:
        # Render non-token table
        current_row = render_non_token_table(
            ws,
            models,
            current_row,
            spec_discounts,
            daily_usages,
            global_discount
        )
    
    # Step 7: Add spacing row after section
    current_row += 1
    
    # Step 8: Return next row number
    return current_row


def group_models_by_category(
    selected_models: List[Dict[str, Any]],
    model_configs: Dict[str, Any]
) -> Dict[str, Dict[str, Any]]:
    """
    Group models by category for rendering.
    
    Args:
        selected_models: List of selected models from the frontend
        model_configs: Dictionary of model configurations keyed by model_id or model_code
    
    Returns:
        Dictionary of grouped models by category:
        {
            'category_key': {
                'config': {...},  # Category configuration
                'models': [
                    {
                        'model': {...},  # Original model data
                        'specs': [...]   # List of normalized spec dicts
                    }
                ]
            }
        }
    
    Logic:
        1. Initialize empty dict for grouped models
        2. For each model in selectedModels:
           - Classify model using classify_model()
           - Get model config (support model_code and id keys)
           - Get specs (support variants, specs, spec fields)
           - For each spec:
             - Extract prices
             - Create normalized spec dict
             - Add to grouped_models[category_key]
        3. Return grouped_models dict
    
    Requirements: 7.1, 7.2, 7.3, 8.1, 8.2, 8.3, 8.4
    
    Examples:
        >>> selected_models = [
        ...     {'id': 1, 'model_code': 'qwen-max', 'model_name': 'Qwen Max'}
        ... ]
        >>> model_configs = {
        ...     '1': {
        ...         'variants': [
        ...             {'id': 1, 'model_name': 'qwen-max', 'mode': 'thinking', 
        ...              'token_tier': '0-128K', 'prices': [...]}
        ...         ]
        ...     }
        ... }
        >>> grouped = group_models_by_category(selected_models, model_configs)
        >>> # Returns: {'text_qwen': {'config': {...}, 'models': [...]}}
    """
    # Step 1: Initialize empty dict for grouped models
    grouped_models = {}
    
    # DEBUG: Log the input data structure
    logger.info(f"=== DEBUG group_models_by_category ===")
    logger.info(f"selected_models count: {len(selected_models)}")
    logger.info(f"model_configs keys: {list(model_configs.keys())}")
    if selected_models:
        logger.info(f"First selected_model sample: {selected_models[0]}")
    if model_configs:
        first_key = list(model_configs.keys())[0]
        logger.info(f"First model_config sample (key={first_key}): {model_configs[first_key]}")
    
    # Step 2: Process each model in selectedModels
    for model in selected_models:
        try:
            # Classify model using classify_model()
            category_key = classify_model(model)
            
            # Get model identifier - try multiple field names
            # Support both 'id' and 'model_code' as keys
            model_id = str(model.get('id') or model.get('model_code') or model.get('name') or '')
            
            if not model_id:
                logger.warning(f"Model has no valid identifier, skipping: {model}")
                continue
            
            # Get model config - try both model_id and model_code as keys
            model_config = None
            
            # Try using model_id as key
            if model_id in model_configs:
                model_config = model_configs[model_id]
            
            # Try using model_code as key
            if model_config is None:
                model_code = model.get('model_code')
                if model_code and model_code in model_configs:
                    model_config = model_configs[model_code]
            
            # If still no config found, skip this model
            if model_config is None:
                logger.debug(f"No config found for model {model_id}, skipping")
                continue
            
            # Get specs - support variants (new), specs (legacy multi-select), and spec (legacy single-select) fields
            specs_list = []
            
            # Try 'variants' field (new format)
            if 'variants' in model_config and model_config['variants']:
                specs_list = model_config['variants']
            
            # Try 'specs' field (legacy multi-select format)
            elif 'specs' in model_config and model_config['specs']:
                specs_list = model_config['specs']
            
            # Try 'spec' field (legacy single-select format)
            elif 'spec' in model_config and model_config['spec']:
                # Wrap single spec in a list
                specs_list = [model_config['spec']]
            
            # If no specs found, skip this model
            if not specs_list:
                logger.debug(f"No specs found for model {model_id}, skipping")
                continue
            
            # Process each spec
            normalized_specs = []
            for spec in specs_list:
                if not isinstance(spec, dict):
                    continue
                
                # DEBUG: Log spec data before extraction
                logger.info(f"Processing spec for model {model_id}: {spec}")
                
                # Extract prices using extract_prices()
                price_data = extract_prices(spec)
                
                # DEBUG: Log extracted price data
                logger.info(f"Extracted price_data: {price_data}")
                
                # Create normalized spec dict
                normalized_spec = {
                    'id': spec.get('id') or spec.get('spec_id'),
                    'model_name': spec.get('model_name') or model.get('model_name') or model.get('name'),
                    'mode': spec.get('mode') or '-',
                    # Support both token_tier (new) and token_range (legacy) fields
                    'token_range': spec.get('token_tier') or spec.get('token_range') or '-',
                    'input_price': price_data['input_price'],
                    'output_price': price_data['output_price'],
                    'non_token_price': price_data['non_token_price'],
                    'price_unit': price_data['price_unit'],
                    'dimension_code': price_data['dimension_code'],
                    'remark': spec.get('remark') or ''
                }
                
                normalized_specs.append(normalized_spec)
            
            # Skip if no valid specs were normalized
            if not normalized_specs:
                logger.debug(f"No valid specs for model {model_id}, skipping")
                continue
            
            # Initialize category in grouped_models if not exists
            if category_key not in grouped_models:
                category_config = get_category_config(category_key)
                if category_config is None:
                    logger.warning(f"Invalid category key: {category_key}, using default")
                    category_config = {
                        'name': category_key,
                        'icon': '📋',
                        'price_type': 'token',
                        'order': 999
                    }
                
                grouped_models[category_key] = {
                    'config': category_config,
                    'models': []
                }
            
            # Add model with normalized specs to grouped_models[category_key]
            grouped_models[category_key]['models'].append({
                'model': model,
                'specs': normalized_specs
            })
        
        except Exception as e:
            logger.error(f"Error processing model {model.get('id') or model.get('model_code')}: {e}")
            # Continue processing other models
            continue
    
    # Step 3: Return grouped_models dict
    return grouped_models


def render_non_token_table(
    ws,
    models: List[Dict[str, Any]],
    start_row: int,
    spec_discounts: Dict[str, Dict[str, float]],
    daily_usages: Dict[str, Dict[str, str]],
    global_discount: float
) -> int:
    """
    Render table for non-token models (image generation, TTS, ASR, video generation).
    
    Args:
        ws: Excel worksheet (openpyxl Worksheet object)
        models: List of model data with specs to render
        start_row: Starting row number for the table
        spec_discounts: Nested dict of spec-level discounts {model_id: {spec_id: discount_percent}}
        daily_usages: Nested dict of daily usage values {model_id: {spec_id: usage_value}}
        global_discount: Global discount percentage (0-100)
    
    Returns:
        Next available row number after the table
    
    Columns (without discounts):
        序号, 模型名称, 单价, 单位, 日估计用量, 预估月用量, 预估月费
    
    Columns (with discounts):
        序号, 模型名称, 单价, 单位, 折扣, 折后单价, 日估计用量, 预估月用量, 预估月费
    
    Logic:
        1. Determine if any discounts exist
        2. Render appropriate headers
        3. For each model and spec:
           - Extract non-token price and unit
           - Get spec discount
           - Calculate discounted price if applicable
           - Get daily usage
           - Calculate monthly usage and cost
           - Render data row
    
    Requirements: 1.2, 5.4, 9.2, 9.3
    """
    current_row = start_row
    
    # Step 1: Determine if any discounts exist
    has_discount = has_any_discount(spec_discounts, global_discount)
    
    # Step 2: Render table headers
    if has_discount:
        # Headers with discount columns
        headers = [
            '序号', '模型名称', '单价', '单位', '折扣', '折后单价',
            '日估计用量', '预估月用量', '预估月费'
        ]
    else:
        # Headers without discount columns
        headers = [
            '序号', '模型名称', '单价', '单位',
            '日估计用量', '预估月用量', '预估月费'
        ]
    
    # Apply header styling: bold, white text, blue background #4472C4, center alignment, borders
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    current_row += 1
    
    # Step 3: Render data rows for each model and spec
    row_number = 1
    
    for model_data in models:
        model = model_data.get('model', {})
        specs = model_data.get('specs', [])
        
        # Get model identifier (try multiple field names)
        model_id = str(model.get('id') or model.get('model_code') or model.get('name') or '')
        model_name = model.get('model_name') or model.get('name') or model.get('model_code') or '-'
        
        for spec in specs:
            # Get spec identifier
            spec_id = str(spec.get('id') or spec.get('spec_id') or '')
            
            # Check if spec is already normalized (has non_token_price field)
            # or if we need to extract prices from prices array
            if 'non_token_price' in spec and spec.get('non_token_price') is not None:
                # Use already extracted prices from normalized spec
                non_token_price = spec.get('non_token_price')
                price_unit = spec.get('price_unit') or '-'
            else:
                # Extract prices from prices array (for direct calls or legacy data)
                price_data = extract_prices(spec)
                non_token_price = price_data['non_token_price']
                price_unit = price_data['price_unit'] or '-'
            
            # Get spec discount using get_spec_discount()
            discount = get_spec_discount(model_id, spec_id, spec_discounts, global_discount)
            
            # Calculate discounted price if applicable
            discount_rate = (100 - discount) / 100
            discounted_price = non_token_price * discount_rate if non_token_price is not None else None
            
            # Get daily usage from daily_usages dict
            daily_usage = ''
            if model_id in daily_usages and spec_id in daily_usages[model_id]:
                daily_usage = daily_usages[model_id][spec_id]
            
            # Calculate monthly usage
            monthly_usage = calculate_monthly_usage(daily_usage)
            
            # Calculate monthly cost
            # For non-token models, we need to construct price_data dict
            # Check if we already have extracted prices or need to extract them
            if 'non_token_price' in spec and spec.get('non_token_price') is not None:
                # Use already extracted prices
                price_data_for_calc = {
                    'input_price': spec.get('input_price'),
                    'output_price': spec.get('output_price'),
                    'non_token_price': non_token_price,
                    'price_unit': price_unit,
                    'dimension_code': spec.get('dimension_code')
                }
            else:
                # Extract prices (price_data was already extracted above)
                price_data_for_calc = {
                    'input_price': None,
                    'output_price': None,
                    'non_token_price': non_token_price,
                    'price_unit': price_unit,
                    'dimension_code': None
                }
            
            monthly_cost = calculate_monthly_cost(
                price_data_for_calc,
                daily_usage,
                discount,
                'non_token'  # Use 'non_token' as price_type for non-token models
            )
            
            # Render data row
            col_idx = 1
            
            # 序号
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = row_number
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # 模型名称
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = model_name
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # 单价
            cell = ws.cell(row=current_row, column=col_idx)
            if non_token_price is not None:
                cell.value = f"¥{non_token_price:.4f}"
            else:
                cell.value = '-'
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # 单位
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = price_unit
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # Discount columns (if applicable)
            if has_discount:
                # 折扣
                cell = ws.cell(row=current_row, column=col_idx)
                if discount > 0:
                    cell.value = f"{discount:.1f}%"
                else:
                    cell.value = '-'
                cell.font = Font(name='微软雅黑', size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                col_idx += 1
                
                # 折后单价
                cell = ws.cell(row=current_row, column=col_idx)
                if discounted_price is not None:
                    cell.value = f"¥{discounted_price:.4f}"
                else:
                    cell.value = '-'
                cell.font = Font(name='微软雅黑', size=10)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                col_idx += 1
            
            # 日估计用量
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = daily_usage if daily_usage else '-'
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # 预估月用量
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = monthly_usage
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            col_idx += 1
            
            # 预估月费
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = monthly_cost
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            current_row += 1
            row_number += 1
    
    # Return next row number
    return current_row


class ExcelExporter:
    """Excel导出器"""
    
    def __init__(self):
        self.title_font = Font(name='微软雅黑', size=16, bold=True)
        self.header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
        self.normal_font = Font(name='微软雅黑', size=10)
        
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    async def generate_standard_quote(
        self,
        quote: QuoteSheet,
        items: List[QuoteItem]
    ) -> bytes:
        """
        生成标准报价单
        
        Args:
            quote: 报价单主记录
            items: 报价明细列表
        
        Returns:
            Excel文件字节流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "报价单"
        
        # 设置列宽
        column_widths = {
            'A': 5,   # 序号
            'B': 25,  # 产品名称
            'C': 20,  # 规格配置
            'D': 10,  # 数量
            'E': 10,  # 时长
            'F': 15,  # 单价
            'G': 15,  # 小计
            'H': 20   # 备注
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 1. 标题部分
        current_row = 1
        ws.merge_cells(f'A{current_row}:H{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "阿里云产品报价单"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 2
        
        # 2. 基本信息
        info_data = [
            ['客户名称', quote.customer_name, '项目名称', quote.project_name or '-'],
            ['报价日期', quote.created_at.strftime('%Y-%m-%d'), '有效期至', 
             quote.valid_until.strftime('%Y-%m-%d') if quote.valid_until else '-'],
            ['币种', quote.currency, '报价单号', str(quote.quote_id)[:8].upper()]
        ]
        
        for row_data in info_data:
            ws[f'A{current_row}'] = row_data[0]
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'] = row_data[1]
            ws[f'D{current_row}'] = row_data[2]
            ws[f'D{current_row}'].font = Font(bold=True)
            ws[f'E{current_row}'] = row_data[3]
            current_row += 1
        
        current_row += 1
        
        # 3. 表头
        headers = ['序号', '产品名称', '规格配置', '数量', '时长(月)', '单价(元)', '小计(元)', '备注']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        current_row += 1
        data_start_row = current_row
        
        # 4. 数据行
        for idx, item in enumerate(items, 1):
            # 序号
            ws.cell(row=current_row, column=1, value=idx)
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
            
            # 产品名称
            ws.cell(row=current_row, column=2, value=item.product_name)
            
            # 规格配置
            spec_text = self._format_spec_config(item.spec_config)
            ws.cell(row=current_row, column=3, value=spec_text)
            
            # 数量
            ws.cell(row=current_row, column=4, value=item.quantity)
            ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center')
            
            # 时长
            ws.cell(row=current_row, column=5, value=item.duration_months or '-')
            ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='center')
            
            # 单价
            ws.cell(row=current_row, column=6, value=float(item.unit_price))
            ws.cell(row=current_row, column=6).number_format = '#,##0.00'
            
            # 小计
            ws.cell(row=current_row, column=7, value=float(item.subtotal))
            ws.cell(row=current_row, column=7).number_format = '#,##0.00'
            
            # 备注(折扣信息)
            remark = self._format_discount_info(item.discount_info)
            ws.cell(row=current_row, column=8, value=remark)
            
            # 应用边框
            for col_idx in range(1, 9):
                ws.cell(row=current_row, column=col_idx).border = self.thin_border
            
            current_row += 1
        
        # 5. 合计行
        ws.merge_cells(f'A{current_row}:F{current_row}')
        total_cell = ws[f'A{current_row}']
        total_cell.value = "报价总计"
        total_cell.font = Font(bold=True, size=11)
        total_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_cell.fill = self.total_fill
        
        total_amount_cell = ws.cell(row=current_row, column=7)
        total_amount_cell.value = float(quote.total_amount)
        total_amount_cell.number_format = '#,##0.00'
        total_amount_cell.font = Font(bold=True, size=11)
        total_amount_cell.fill = self.total_fill
        
        # 应用边框
        for col_idx in range(1, 9):
            ws.cell(row=current_row, column=col_idx).border = self.thin_border
        
        current_row += 2
        
        # 6. 备注说明
        ws.merge_cells(f'A{current_row}:H{current_row}')
        note_cell = ws[f'A{current_row}']
        note_cell.value = "备注: 1. 以上价格为估算价格,实际价格以阿里云官网为准 2. 本报价单有效期30天"
        note_cell.font = Font(size=9, italic=True, color="808080")
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:H{current_row}')
        contact_cell = ws[f'A{current_row}']
        contact_cell.value = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | 报价侠系统"
        contact_cell.font = Font(size=9, italic=True, color="808080")
        
        # 转换为字节流
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def _format_spec_config(self, spec_config: Optional[Dict[str, Any]]) -> str:
        """格式化规格配置"""
        if not spec_config:
            return "-"
        
        parts = []
        for key, value in spec_config.items():
            if key in ['region', 'spec_type', 'billing_mode']:
                continue
            parts.append(f"{key}: {value}")
        
        return "\n".join(parts) if parts else "-"
    
    def _format_discount_info(self, discount_info: Optional[Dict[str, Any]]) -> str:
        """格式化折扣信息"""
        if not discount_info or not discount_info.get('discounts'):
            return "-"
        
        discounts = discount_info.get('discounts', [])
        parts = []
        for discount in discounts:
            discount_type = discount.get('type', '')
            value = discount.get('value', 0)
            
            if discount_type == 'tiered':
                parts.append(f"阶梯折扣: {value}折")
            elif discount_type == 'batch':
                parts.append(f"Batch折扣: {value}折")
            elif discount_type == 'thinking_mode':
                parts.append(f"思考模式: {value}倍")
            elif discount_type == 'package':
                parts.append("套餐价格")
        
        return "\n".join(parts) if parts else "-"
    
    async def generate_competitor_comparison(
        self,
        quote: QuoteSheet,
        items: List[QuoteItem],
        competitor_data: Dict[str, Any]
    ) -> bytes:
        """
        生成竞品对比版报价单
        
        Args:
            quote: 报价单主记录
            items: 报价明细列表
            competitor_data: 竞品数据
        
        Returns:
            Excel文件字节流
        """
        # 竞品对比版本可以后续实现
        # 目前先返回标准版本
        return await self.generate_standard_quote(quote, items)
    
    async def generate_simplified_quote(
        self,
        quote: QuoteSheet,
        items: List[QuoteItem]
    ) -> bytes:
        """
        生成简化版报价单
        
        Args:
            quote: 报价单主记录
            items: 报价明细列表
        
        Returns:
            Excel文件字节流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "简化报价单"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        # 标题
        ws['A1'] = "产品名称"
        ws['B1'] = "数量"
        ws['C1'] = "价格(元)"
        
        # 应用样式
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}1']
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 数据行
        row = 2
        for item in items:
            ws.cell(row=row, column=1, value=item.product_name)
            ws.cell(row=row, column=2, value=item.quantity)
            ws.cell(row=row, column=3, value=float(item.subtotal))
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1
        
        # 合计
        ws.cell(row=row, column=1, value="总计")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3, value=float(quote.total_amount))
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=3).font = Font(bold=True)
        
        # 转换为字节流
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    async def generate_and_upload(
        self,
        quote: QuoteSheet,
        items: List[QuoteItem],
        template_type: str = "standard"
    ) -> Tuple[bytes, Optional[str]]:
        """
        生成Excel并上传到OSS
        
        Args:
            quote: 报价单主记录
            items: 报价明细列表
            template_type: 模板类型 (standard/simplified/competitor)
        
        Returns:
            Tuple[bytes, Optional[str]]: (Excel文件字节流, OSS下载URL)
        """
        # 根据模板类型生成Excel
        if template_type == "simplified":
            excel_bytes = await self.generate_simplified_quote(quote, items)
        elif template_type == "competitor":
            excel_bytes = await self.generate_competitor_comparison(quote, items, {})
        else:
            excel_bytes = await self.generate_standard_quote(quote, items)
        
        # 上传到OSS
        oss_url = None
        try:
            uploader = get_oss_uploader()
            oss_url = await uploader.upload_quote_file(
                file_content=excel_bytes,
                quote_id=str(quote.quote_id),
                file_type="xlsx"
            )
            if oss_url:
                logger.info(f"报价单Excel已上传: {oss_url}")
            else:
                logger.warning("报价单Excel上传失败，OSS可能未配置")
        except Exception as e:
            logger.error(f"上传报价单Excel异常: {e}")
        
        return excel_bytes, oss_url
    
    async def batch_export(
        self,
        quotes_data: List[Tuple[QuoteSheet, List[QuoteItem]]],
        template_type: str = "standard",
        upload_to_oss: bool = True
    ) -> List[Dict[str, Any]]:
        """
        批量导出报价单
        
        Args:
            quotes_data: 报价单数据列表 [(quote, items), ...]
            template_type: 模板类型
            upload_to_oss: 是否上传到OSS
        
        Returns:
            导出结果列表
        """
        results = []
        
        for quote, items in quotes_data:
            try:
                if upload_to_oss:
                    excel_bytes, oss_url = await self.generate_and_upload(
                        quote, items, template_type
                    )
                    results.append({
                        "quote_id": str(quote.quote_id),
                        "quote_no": quote.quote_no,
                        "success": True,
                        "oss_url": oss_url,
                        "file_size": len(excel_bytes)
                    })
                else:
                    if template_type == "simplified":
                        excel_bytes = await self.generate_simplified_quote(quote, items)
                    else:
                        excel_bytes = await self.generate_standard_quote(quote, items)
                    
                    results.append({
                        "quote_id": str(quote.quote_id),
                        "quote_no": quote.quote_no,
                        "success": True,
                        "file_size": len(excel_bytes)
                    })
            except Exception as e:
                logger.error(f"导出报价单 {quote.quote_no} 失败: {e}")
                results.append({
                    "quote_id": str(quote.quote_id),
                    "quote_no": quote.quote_no,
                    "success": False,
                    "error": str(e)
                })
        
        return results


# 全局导出器实例
_exporter: Optional[ExcelExporter] = None


def get_excel_exporter() -> ExcelExporter:
    """获取Excel导出器实例"""
    global _exporter
    if _exporter is None:
        _exporter = ExcelExporter()
    return _exporter

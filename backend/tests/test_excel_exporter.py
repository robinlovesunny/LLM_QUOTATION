"""
Excel导出服务测试
"""
import pytest
from decimal import Decimal
from datetime import datetime, timedelta
from uuid import uuid4
from io import BytesIO
from openpyxl import load_workbook

from app.services.excel_exporter import ExcelExporter, get_excel_exporter
from app.models.quote import QuoteSheet, QuoteItem


class MockQuoteSheet:
    """模拟报价单对象"""
    def __init__(self, **kwargs):
        self.quote_id = kwargs.get('quote_id', uuid4())
        self.quote_no = kwargs.get('quote_no', 'QT202601080001')
        self.customer_name = kwargs.get('customer_name', '测试客户')
        self.project_name = kwargs.get('project_name', '测试项目')
        self.status = kwargs.get('status', 'draft')
        self.total_amount = kwargs.get('total_amount', Decimal('10000.00'))
        self.currency = kwargs.get('currency', 'CNY')
        self.created_at = kwargs.get('created_at', datetime.now())
        self.valid_until = kwargs.get('valid_until', datetime.now() + timedelta(days=30))


class MockQuoteItem:
    """模拟报价项对象"""
    def __init__(self, **kwargs):
        self.item_id = kwargs.get('item_id', uuid4())
        self.product_name = kwargs.get('product_name', 'qwen-max')
        self.spec_config = kwargs.get('spec_config', {'model': 'qwen-max', 'context': '128K'})
        self.quantity = kwargs.get('quantity', 1)
        self.duration_months = kwargs.get('duration_months', 1)
        self.unit_price = kwargs.get('unit_price', Decimal('0.04'))
        self.subtotal = kwargs.get('subtotal', Decimal('1000.00'))
        self.discount_info = kwargs.get('discount_info', None)


class TestExcelExporter:
    """Excel导出服务测试"""
    
    def setup_method(self):
        """初始化测试"""
        self.exporter = ExcelExporter()
    
    @pytest.mark.asyncio
    async def test_generate_standard_quote(self):
        """测试生成标准报价单"""
        # 准备测试数据
        quote = MockQuoteSheet()
        items = [
            MockQuoteItem(
                product_name='qwen-max',
                quantity=1,
                subtotal=Decimal('1000.00')
            ),
            MockQuoteItem(
                product_name='qwen-plus',
                quantity=2,
                subtotal=Decimal('500.00')
            )
        ]
        
        # 生成Excel
        excel_bytes = await self.exporter.generate_standard_quote(quote, items)
        
        # 验证生成的文件
        assert excel_bytes is not None
        assert len(excel_bytes) > 0
        
        # 读取并验证内容
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        
        # 验证标题
        assert ws.title == "报价单"
        assert "阿里云产品报价单" in str(ws['A1'].value)
        
        # 验证客户信息
        customer_found = False
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value == quote.customer_name:
                    customer_found = True
                    break
        assert customer_found, "未找到客户名称"
    
    @pytest.mark.asyncio
    async def test_generate_simplified_quote(self):
        """测试生成简化版报价单"""
        quote = MockQuoteSheet(total_amount=Decimal('5000.00'))
        items = [
            MockQuoteItem(
                product_name='测试产品',
                quantity=5,
                subtotal=Decimal('5000.00')
            )
        ]
        
        # 生成Excel
        excel_bytes = await self.exporter.generate_simplified_quote(quote, items)
        
        # 验证
        assert excel_bytes is not None
        
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        
        assert ws.title == "简化报价单"
        
        # 验证表头
        assert ws['A1'].value == "产品名称"
        assert ws['B1'].value == "数量"
        assert ws['C1'].value == "价格(元)"
        
        # 验证数据行
        assert ws['A2'].value == "测试产品"
        assert ws['B2'].value == 5
    
    @pytest.mark.asyncio
    async def test_empty_items_list(self):
        """测试空商品列表"""
        quote = MockQuoteSheet(total_amount=Decimal('0.00'))
        items = []
        
        # 生成Excel
        excel_bytes = await self.exporter.generate_standard_quote(quote, items)
        
        assert excel_bytes is not None
        assert len(excel_bytes) > 0
    
    @pytest.mark.asyncio
    async def test_competitor_comparison(self):
        """测试竞品对比版（当前返回标准版）"""
        quote = MockQuoteSheet()
        items = [MockQuoteItem()]
        competitor_data = {"competitor": "test"}
        
        excel_bytes = await self.exporter.generate_competitor_comparison(
            quote, items, competitor_data
        )
        
        assert excel_bytes is not None
    
    def test_format_spec_config(self):
        """测试规格配置格式化"""
        # 测试有效配置
        spec_config = {
            'model': 'qwen-max',
            'context': '128K',
            'region': 'cn-beijing',  # 应该被过滤
            'spec_type': 'standard'  # 应该被过滤
        }
        result = self.exporter._format_spec_config(spec_config)
        assert 'model' in result
        assert 'context' in result
        assert 'region' not in result
        
        # 测试空配置
        assert self.exporter._format_spec_config(None) == "-"
        assert self.exporter._format_spec_config({}) == "-"
    
    def test_format_discount_info(self):
        """测试折扣信息格式化"""
        # 测试有效折扣
        discount_info = {
            'discounts': [
                {'type': 'tiered', 'value': 9},
                {'type': 'batch', 'value': 5}
            ]
        }
        result = self.exporter._format_discount_info(discount_info)
        assert '阶梯折扣' in result
        assert 'Batch折扣' in result
        
        # 测试空折扣
        assert self.exporter._format_discount_info(None) == "-"
        assert self.exporter._format_discount_info({}) == "-"
    
    def test_get_excel_exporter_singleton(self):
        """测试获取导出器单例"""
        exporter1 = get_excel_exporter()
        exporter2 = get_excel_exporter()
        
        assert exporter1 is exporter2


class TestExcelExporterEdgeCases:
    """Excel导出器边界情况测试"""
    
    def setup_method(self):
        self.exporter = ExcelExporter()
    
    @pytest.mark.asyncio
    async def test_large_amount(self):
        """测试大金额数值"""
        quote = MockQuoteSheet(total_amount=Decimal('9999999999.99'))
        items = [
            MockQuoteItem(
                subtotal=Decimal('9999999999.99'),
                unit_price=Decimal('99999.99')
            )
        ]
        
        excel_bytes = await self.exporter.generate_standard_quote(quote, items)
        assert excel_bytes is not None
    
    @pytest.mark.asyncio
    async def test_special_characters_in_names(self):
        """测试名称中的特殊字符"""
        quote = MockQuoteSheet(
            customer_name='测试<客户>&"特殊"',
            project_name='项目/名称\\test'
        )
        items = [
            MockQuoteItem(
                product_name='产品<名称>&"测试"'
            )
        ]
        
        excel_bytes = await self.exporter.generate_standard_quote(quote, items)
        assert excel_bytes is not None
    
    @pytest.mark.asyncio
    async def test_many_items(self):
        """测试大量商品（性能测试）"""
        quote = MockQuoteSheet()
        items = [
            MockQuoteItem(
                product_name=f'产品_{i}',
                quantity=i + 1,
                subtotal=Decimal(str(100 * (i + 1)))
            )
            for i in range(100)
        ]
        
        excel_bytes = await self.exporter.generate_standard_quote(quote, items)
        assert excel_bytes is not None
        
        # 验证文件不会太大
        assert len(excel_bytes) < 1024 * 1024  # < 1MB



class TestCategoryConfiguration:
    """Category configuration constants and utilities tests"""
    
    def test_category_config_structure(self):
        """Test CATEGORY_CONFIG has all required categories with correct structure"""
        from app.services.excel_exporter import CATEGORY_CONFIG
        
        # Verify all 12 categories exist
        expected_categories = [
            'text_qwen', 'text_qwen_opensource', 'text_thirdparty',
            'image_gen', 'image_gen_thirdparty', 'tts', 'asr', 'video_gen',
            'text_embedding', 'multimodal_embedding', 'text_nlu', 'industry'
        ]
        
        assert len(CATEGORY_CONFIG) == 12
        for category in expected_categories:
            assert category in CATEGORY_CONFIG
        
        # Verify each category has required fields
        for category_key, config in CATEGORY_CONFIG.items():
            assert 'name' in config, f"Category {category_key} missing 'name'"
            assert 'icon' in config, f"Category {category_key} missing 'icon'"
            assert 'price_type' in config, f"Category {category_key} missing 'price_type'"
            assert 'order' in config, f"Category {category_key} missing 'order'"
            
            # Verify field types
            assert isinstance(config['name'], str)
            assert isinstance(config['icon'], str)
            assert isinstance(config['price_type'], str)
            assert isinstance(config['order'], int)
    
    def test_category_config_icons(self):
        """Test all categories have unique icons"""
        from app.services.excel_exporter import CATEGORY_CONFIG
        
        expected_icons = {
            'text_qwen': '💬',
            'text_qwen_opensource': '📝',
            'text_thirdparty': '🤖',
            'image_gen': '🎨',
            'image_gen_thirdparty': '🖼️',
            'tts': '🔊',
            'asr': '🎤',
            'video_gen': '🎬',
            'text_embedding': '📊',
            'multimodal_embedding': '🌐',
            'text_nlu': '🔍',
            'industry': '🏭'
        }
        
        for category_key, expected_icon in expected_icons.items():
            assert CATEGORY_CONFIG[category_key]['icon'] == expected_icon
    
    def test_category_config_price_types(self):
        """Test categories have correct price types"""
        from app.services.excel_exporter import CATEGORY_CONFIG
        
        # Token-based categories
        token_categories = [
            'text_qwen', 'text_qwen_opensource', 'text_thirdparty',
            'text_embedding', 'multimodal_embedding', 'text_nlu', 'industry'
        ]
        for category in token_categories:
            assert CATEGORY_CONFIG[category]['price_type'] == 'token'
        
        # Non-token categories
        assert CATEGORY_CONFIG['image_gen']['price_type'] == 'image'
        assert CATEGORY_CONFIG['image_gen_thirdparty']['price_type'] == 'image'
        assert CATEGORY_CONFIG['tts']['price_type'] == 'character'
        assert CATEGORY_CONFIG['asr']['price_type'] == 'audio'
        assert CATEGORY_CONFIG['video_gen']['price_type'] == 'video'
    
    def test_category_config_ordering(self):
        """Test categories have correct order values (1-12)"""
        from app.services.excel_exporter import CATEGORY_CONFIG
        
        orders = [config['order'] for config in CATEGORY_CONFIG.values()]
        
        # Verify all orders are unique and sequential from 1 to 12
        assert sorted(orders) == list(range(1, 13))
        
        # Verify specific order assignments
        assert CATEGORY_CONFIG['text_qwen']['order'] == 1
        assert CATEGORY_CONFIG['text_qwen_opensource']['order'] == 2
        assert CATEGORY_CONFIG['text_thirdparty']['order'] == 3
        assert CATEGORY_CONFIG['image_gen']['order'] == 4
        assert CATEGORY_CONFIG['image_gen_thirdparty']['order'] == 5
        assert CATEGORY_CONFIG['tts']['order'] == 6
        assert CATEGORY_CONFIG['asr']['order'] == 7
        assert CATEGORY_CONFIG['video_gen']['order'] == 8
        assert CATEGORY_CONFIG['text_embedding']['order'] == 9
        assert CATEGORY_CONFIG['multimodal_embedding']['order'] == 10
        assert CATEGORY_CONFIG['text_nlu']['order'] == 11
        assert CATEGORY_CONFIG['industry']['order'] == 12
    
    def test_unit_map_structure(self):
        """Test UNIT_MAP has all required dimension codes"""
        from app.services.excel_exporter import UNIT_MAP
        
        # Verify all 4 dimension codes exist
        expected_codes = ['character', 'audio_second', 'video_second', 'image_count']
        
        assert len(UNIT_MAP) == 4
        for code in expected_codes:
            assert code in UNIT_MAP
        
        # Verify correct mappings
        assert UNIT_MAP['character'] == '字符'
        assert UNIT_MAP['audio_second'] == '秒'
        assert UNIT_MAP['video_second'] == '秒'
        assert UNIT_MAP['image_count'] == '张'
    
    def test_get_category_config_valid(self):
        """Test get_category_config returns correct config for valid keys"""
        from app.services.excel_exporter import get_category_config
        
        # Test valid category keys
        config = get_category_config('text_qwen')
        assert config is not None
        assert config['name'] == '文本生成-通义千问'
        assert config['icon'] == '💬'
        assert config['price_type'] == 'token'
        assert config['order'] == 1
        
        config = get_category_config('image_gen')
        assert config is not None
        assert config['name'] == '图像生成'
        assert config['icon'] == '🎨'
        assert config['price_type'] == 'image'
        assert config['order'] == 4
    
    def test_get_category_config_invalid(self):
        """Test get_category_config returns None for invalid keys"""
        from app.services.excel_exporter import get_category_config
        
        # Test invalid category keys
        assert get_category_config('invalid_category') is None
        assert get_category_config('') is None
        assert get_category_config('text') is None
        assert get_category_config('image') is None
    
    def test_get_category_config_all_categories(self):
        """Test get_category_config works for all 12 categories"""
        from app.services.excel_exporter import get_category_config, CATEGORY_CONFIG
        
        for category_key in CATEGORY_CONFIG.keys():
            config = get_category_config(category_key)
            assert config is not None
            assert config == CATEGORY_CONFIG[category_key]


class TestModelClassification:
    """Model classification tests - Requirements 7.1, 7.2, 7.3"""
    
    def test_classify_by_category_field(self):
        """Test classification using category field (Requirement 7.1)"""
        from app.services.excel_exporter import classify_model
        
        # Test with exact category match
        model = {'category': 'text_qwen'}
        assert classify_model(model) == 'text_qwen'
        
        model = {'category': 'image_gen'}
        assert classify_model(model) == 'image_gen'
        
        model = {'category': 'tts'}
        assert classify_model(model) == 'tts'
        
        # Test with uppercase (should be case-insensitive)
        model = {'category': 'TEXT_QWEN'}
        assert classify_model(model) == 'text_qwen'
        
        model = {'category': 'Image_Gen'}
        assert classify_model(model) == 'image_gen'
    
    def test_classify_by_sub_category_field(self):
        """Test classification using sub_category field (Requirement 7.1)"""
        from app.services.excel_exporter import classify_model
        
        # Test with sub_category match
        model = {'sub_category': 'video_gen'}
        assert classify_model(model) == 'video_gen'
        
        model = {'sub_category': 'asr'}
        assert classify_model(model) == 'asr'
        
        # Test sub_category takes precedence when both exist
        model = {'category': 'text_qwen', 'sub_category': 'image_gen'}
        assert classify_model(model) == 'text_qwen'  # category checked first
    
    def test_classify_by_name_pattern_image_gen(self):
        """Test classification by image generation name patterns (Requirement 7.2)"""
        from app.services.excel_exporter import classify_model
        
        # Test wanx pattern
        model = {'model_code': 'wanx-v1'}
        assert classify_model(model) == 'image_gen'
        
        # Test flux pattern
        model = {'model_code': 'flux-schnell'}
        assert classify_model(model) == 'image_gen'
        
        # Test stable-diffusion pattern
        model = {'model_code': 'stable-diffusion-xl'}
        assert classify_model(model) == 'image_gen'
        
        # Test qwen-image pattern
        model = {'name': 'qwen-image-v1'}
        assert classify_model(model) == 'image_gen'
        
        # Test image-edit pattern
        model = {'model_name': 'image-edit-pro'}
        assert classify_model(model) == 'image_gen'
    
    def test_classify_by_name_pattern_video_gen(self):
        """Test classification by video generation name patterns (Requirement 7.2)"""
        from app.services.excel_exporter import classify_model
        
        # Test t2v pattern
        model = {'model_code': 't2v-model'}
        assert classify_model(model) == 'video_gen'
        
        # Test i2v pattern
        model = {'model_code': 'i2v-generator'}
        assert classify_model(model) == 'video_gen'
        
        # Test wan2* pattern
        model = {'model_code': 'wan2video'}
        assert classify_model(model) == 'video_gen'
        
        model = {'name': 'wan2clip'}
        assert classify_model(model) == 'video_gen'
    
    def test_classify_by_name_pattern_tts(self):
        """Test classification by TTS name patterns (Requirement 7.2)"""
        from app.services.excel_exporter import classify_model
        
        # Test -tts suffix
        model = {'model_code': 'sambert-tts'}
        assert classify_model(model) == 'tts'
        
        model = {'model_code': 'azure-tts'}
        assert classify_model(model) == 'tts'
        
        # Test cosyvoice pattern
        model = {'name': 'cosyvoice-v1'}
        assert classify_model(model) == 'tts'
        
        model = {'model_name': 'cosyvoice-pro'}
        assert classify_model(model) == 'tts'
    
    def test_classify_by_name_pattern_asr(self):
        """Test classification by ASR name patterns (Requirement 7.2)"""
        from app.services.excel_exporter import classify_model
        
        # Test -asr suffix
        model = {'model_code': 'whisper-asr'}
        assert classify_model(model) == 'asr'
        
        # Test paraformer pattern
        model = {'model_code': 'paraformer-v1'}
        assert classify_model(model) == 'asr'
        
        # Test sensevoice pattern
        model = {'name': 'sensevoice-small'}
        assert classify_model(model) == 'asr'
    
    def test_classify_by_name_pattern_embedding(self):
        """Test classification by embedding name patterns (Requirement 7.2)"""
        from app.services.excel_exporter import classify_model
        
        # Test embedding pattern
        model = {'model_code': 'text-embedding-v1'}
        assert classify_model(model) == 'text_embedding'
        
        model = {'name': 'embedding-model'}
        assert classify_model(model) == 'text_embedding'
        
        model = {'model_name': 'qwen-embedding'}
        assert classify_model(model) == 'text_embedding'
    
    def test_classify_default_to_text_qwen(self):
        """Test default classification to text_qwen (Requirement 7.3)"""
        from app.services.excel_exporter import classify_model
        
        # Test with unrecognized model name
        model = {'model_code': 'qwen-max'}
        assert classify_model(model) == 'text_qwen'
        
        model = {'name': 'gpt-4'}
        assert classify_model(model) == 'text_qwen'
        
        # Test with empty model
        model = {}
        assert classify_model(model) == 'text_qwen'
        
        # Test with only irrelevant fields
        model = {'id': 123, 'price': 0.04}
        assert classify_model(model) == 'text_qwen'
    
    def test_classify_field_precedence(self):
        """Test that category field takes precedence over name patterns"""
        from app.services.excel_exporter import classify_model
        
        # Category field should override name pattern
        model = {'category': 'text_qwen', 'model_code': 'wanx-v1'}
        assert classify_model(model) == 'text_qwen'
        
        model = {'sub_category': 'tts', 'name': 'embedding-model'}
        assert classify_model(model) == 'tts'
    
    def test_classify_case_insensitivity(self):
        """Test that classification is case-insensitive"""
        from app.services.excel_exporter import classify_model
        
        # Test uppercase model names
        model = {'model_code': 'WANX-V1'}
        assert classify_model(model) == 'image_gen'
        
        model = {'model_code': 'COSYVOICE'}
        assert classify_model(model) == 'tts'
        
        # Test mixed case
        model = {'name': 'Paraformer-V1'}
        assert classify_model(model) == 'asr'
    
    def test_classify_with_multiple_name_fields(self):
        """Test classification when multiple name fields are present"""
        from app.services.excel_exporter import classify_model
        
        # model_code should be checked first
        model = {
            'model_code': 'wanx-v1',
            'name': 'qwen-max',
            'model_name': 'text-model'
        }
        assert classify_model(model) == 'image_gen'
        
        # name should be checked if model_code is empty
        model = {
            'model_code': '',
            'name': 'cosyvoice',
            'model_name': 'text-model'
        }
        assert classify_model(model) == 'tts'
        
        # model_name should be checked if both model_code and name are empty
        model = {
            'model_code': '',
            'name': '',
            'model_name': 'paraformer'
        }
        assert classify_model(model) == 'asr'
    
    def test_classify_edge_cases(self):
        """Test edge cases in classification"""
        from app.services.excel_exporter import classify_model
        
        # Test with None values
        model = {'category': None, 'model_code': 'wanx'}
        assert classify_model(model) == 'image_gen'
        
        # Test with special characters
        model = {'model_code': 'wanx-v1.0-beta'}
        assert classify_model(model) == 'image_gen'
        
        # Test with spaces
        model = {'name': 'text embedding model'}
        assert classify_model(model) == 'text_embedding'
        
        # Test partial matches - models ending with -tts
        model = {'model_code': 'sambert-zhichu-tts'}
        assert classify_model(model) == 'tts'
        
        model = {'model_code': 'custom-voice-asr'}
        assert classify_model(model) == 'asr'


class TestPriceExtraction:
    """Price extraction tests - Requirements 3.1, 3.2, 3.3, 3.4, 3.5"""
    
    def test_extract_input_prices_from_prices_array(self):
        """Test extracting input prices from dimension_codes: input, input_token, input_token_image (Requirement 3.1)"""
        from app.services.excel_exporter import extract_prices
        
        # Test 'input' dimension_code
        spec = {
            'prices': [
                {'dimension_code': 'input', 'unit_price': 0.04}
            ]
        }
        result = extract_prices(spec)
        assert result['input_price'] == 0.04
        assert result['output_price'] is None
        assert result['non_token_price'] is None
        
        # Test 'input_token' dimension_code
        spec = {
            'prices': [
                {'dimension_code': 'input_token', 'unit_price': 0.05}
            ]
        }
        result = extract_prices(spec)
        assert result['input_price'] == 0.05
        
        # Test 'input_token_image' dimension_code
        spec = {
            'prices': [
                {'dimension_code': 'input_token_image', 'unit_price': 0.06}
            ]
        }
        result = extract_prices(spec)
        assert result['input_price'] == 0.06
    
    def test_extract_output_prices_from_prices_array(self):
        """Test extracting output prices from dimension_codes: output, output_token, output_token_thinking (Requirement 3.2)"""
        from app.services.excel_exporter import extract_prices
        
        # Test 'output' dimension_code
        spec = {
            'prices': [
                {'dimension_code': 'output', 'unit_price': 0.12}
            ]
        }
        result = extract_prices(spec)
        assert result['output_price'] == 0.12
        assert result['input_price'] is None
        assert result['non_token_price'] is None
        
        # Test 'output_token' dimension_code
        spec = {
            'prices': [
                {'dimension_code': 'output_token', 'unit_price': 0.13}
            ]
        }
        result = extract_prices(spec)
        assert result['output_price'] == 0.13
        
        # Test 'output_token_thinking' dimension_code
        spec = {
            'prices': [
                {'dimension_code': 'output_token_thinking', 'unit_price': 0.14}
            ]
        }
        result = extract_prices(spec)
        assert result['output_price'] == 0.14
    
    def test_extract_non_token_prices_from_prices_array(self):
        """Test extracting non-token prices from dimension_codes: character, audio_second, video_second, image_count (Requirement 3.3)"""
        from app.services.excel_exporter import extract_prices
        
        # Test 'character' dimension_code
        spec = {
            'prices': [
                {'dimension_code': 'character', 'unit_price': 0.001}
            ]
        }
        result = extract_prices(spec)
        assert result['non_token_price'] == 0.001
        assert result['dimension_code'] == 'character'
        assert result['input_price'] is None
        assert result['output_price'] is None
        
        # Test 'audio_second' dimension_code
        spec = {
            'prices': [
                {'dimension_code': 'audio_second', 'unit_price': 0.002}
            ]
        }
        result = extract_prices(spec)
        assert result['non_token_price'] == 0.002
        assert result['dimension_code'] == 'audio_second'
        
        # Test 'video_second' dimension_code
        spec = {
            'prices': [
                {'dimension_code': 'video_second', 'unit_price': 0.5}
            ]
        }
        result = extract_prices(spec)
        assert result['non_token_price'] == 0.5
        assert result['dimension_code'] == 'video_second'
        
        # Test 'image_count' dimension_code
        spec = {
            'prices': [
                {'dimension_code': 'image_count', 'unit_price': 0.08}
            ]
        }
        result = extract_prices(spec)
        assert result['non_token_price'] == 0.08
        assert result['dimension_code'] == 'image_count'
    
    def test_map_dimension_code_to_unit(self):
        """Test mapping dimension_code to unit labels using UNIT_MAP (Requirement 3.4)"""
        from app.services.excel_exporter import extract_prices
        
        # Test character → 字符
        spec = {
            'prices': [
                {'dimension_code': 'character', 'unit_price': 0.001}
            ]
        }
        result = extract_prices(spec)
        assert result['price_unit'] == '字符'
        
        # Test audio_second → 秒
        spec = {
            'prices': [
                {'dimension_code': 'audio_second', 'unit_price': 0.002}
            ]
        }
        result = extract_prices(spec)
        assert result['price_unit'] == '秒'
        
        # Test video_second → 秒
        spec = {
            'prices': [
                {'dimension_code': 'video_second', 'unit_price': 0.5}
            ]
        }
        result = extract_prices(spec)
        assert result['price_unit'] == '秒'
        
        # Test image_count → 张
        spec = {
            'prices': [
                {'dimension_code': 'image_count', 'unit_price': 0.08}
            ]
        }
        result = extract_prices(spec)
        assert result['price_unit'] == '张'
    
    def test_extract_multiple_prices_from_array(self):
        """Test extracting both input and output prices from same array"""
        from app.services.excel_exporter import extract_prices
        
        # Test with both input and output
        spec = {
            'prices': [
                {'dimension_code': 'input', 'unit_price': 0.04},
                {'dimension_code': 'output', 'unit_price': 0.12}
            ]
        }
        result = extract_prices(spec)
        assert result['input_price'] == 0.04
        assert result['output_price'] == 0.12
        assert result['non_token_price'] is None
        
        # Test with multiple input types (last one should win)
        spec = {
            'prices': [
                {'dimension_code': 'input', 'unit_price': 0.04},
                {'dimension_code': 'input_token', 'unit_price': 0.05},
                {'dimension_code': 'input_token_image', 'unit_price': 0.06}
            ]
        }
        result = extract_prices(spec)
        assert result['input_price'] == 0.06  # Last one wins
    
    def test_fallback_to_legacy_fields(self):
        """Test fallback to legacy input_price/output_price fields when prices array missing (Requirement 3.5)"""
        from app.services.excel_exporter import extract_prices
        
        # Test with legacy fields only
        spec = {
            'input_price': 0.04,
            'output_price': 0.12
        }
        result = extract_prices(spec)
        assert result['input_price'] == 0.04
        assert result['output_price'] == 0.12
        assert result['non_token_price'] is None
        
        # Test with only input_price
        spec = {
            'input_price': 0.05
        }
        result = extract_prices(spec)
        assert result['input_price'] == 0.05
        assert result['output_price'] is None
        
        # Test with only output_price
        spec = {
            'output_price': 0.13
        }
        result = extract_prices(spec)
        assert result['input_price'] is None
        assert result['output_price'] == 0.13
    
    def test_prices_array_takes_precedence_over_legacy(self):
        """Test that prices array takes precedence over legacy fields"""
        from app.services.excel_exporter import extract_prices
        
        # When both exist, prices array should be used
        spec = {
            'prices': [
                {'dimension_code': 'input', 'unit_price': 0.10},
                {'dimension_code': 'output', 'unit_price': 0.20}
            ],
            'input_price': 0.04,
            'output_price': 0.12
        }
        result = extract_prices(spec)
        assert result['input_price'] == 0.10  # From prices array
        assert result['output_price'] == 0.20  # From prices array
    
    def test_extract_prices_with_empty_prices_array(self):
        """Test with empty prices array falls back to legacy"""
        from app.services.excel_exporter import extract_prices
        
        # Empty prices array should trigger fallback
        spec = {
            'prices': [],
            'input_price': 0.04,
            'output_price': 0.12
        }
        result = extract_prices(spec)
        assert result['input_price'] == 0.04
        assert result['output_price'] == 0.12
    
    def test_extract_prices_with_invalid_price_entries(self):
        """Test handling of invalid price entries in array"""
        from app.services.excel_exporter import extract_prices
        
        # Test with missing unit_price
        spec = {
            'prices': [
                {'dimension_code': 'input'},  # Missing unit_price
                {'dimension_code': 'output', 'unit_price': 0.12}
            ]
        }
        result = extract_prices(spec)
        assert result['input_price'] is None  # Skipped invalid entry
        assert result['output_price'] == 0.12
        
        # Test with None unit_price
        spec = {
            'prices': [
                {'dimension_code': 'input', 'unit_price': None},
                {'dimension_code': 'output', 'unit_price': 0.12}
            ]
        }
        result = extract_prices(spec)
        assert result['input_price'] is None
        assert result['output_price'] == 0.12
        
        # Test with non-dict entries
        spec = {
            'prices': [
                'invalid',
                {'dimension_code': 'output', 'unit_price': 0.12}
            ]
        }
        result = extract_prices(spec)
        assert result['output_price'] == 0.12
    
    def test_extract_prices_with_string_prices(self):
        """Test conversion of string prices to float"""
        from app.services.excel_exporter import extract_prices
        
        # Test with string prices
        spec = {
            'prices': [
                {'dimension_code': 'input', 'unit_price': '0.04'},
                {'dimension_code': 'output', 'unit_price': '0.12'}
            ]
        }
        result = extract_prices(spec)
        assert result['input_price'] == 0.04
        assert result['output_price'] == 0.12
        
        # Test with legacy string prices
        spec = {
            'input_price': '0.05',
            'output_price': '0.13'
        }
        result = extract_prices(spec)
        assert result['input_price'] == 0.05
        assert result['output_price'] == 0.13
    
    def test_extract_prices_with_invalid_string_prices(self):
        """Test handling of invalid string prices"""
        from app.services.excel_exporter import extract_prices
        
        # Test with invalid string in prices array
        spec = {
            'prices': [
                {'dimension_code': 'input', 'unit_price': 'invalid'},
                {'dimension_code': 'output', 'unit_price': 0.12}
            ]
        }
        result = extract_prices(spec)
        assert result['input_price'] is None  # Invalid conversion skipped
        assert result['output_price'] == 0.12
        
        # Test with invalid legacy prices
        spec = {
            'input_price': 'invalid',
            'output_price': 0.12
        }
        result = extract_prices(spec)
        assert result['input_price'] is None
        assert result['output_price'] == 0.12
    
    def test_extract_prices_case_insensitivity(self):
        """Test that dimension_code matching is case-insensitive"""
        from app.services.excel_exporter import extract_prices
        
        # Test with uppercase dimension_codes
        spec = {
            'prices': [
                {'dimension_code': 'INPUT', 'unit_price': 0.04},
                {'dimension_code': 'OUTPUT', 'unit_price': 0.12}
            ]
        }
        result = extract_prices(spec)
        assert result['input_price'] == 0.04
        assert result['output_price'] == 0.12
        
        # Test with mixed case
        spec = {
            'prices': [
                {'dimension_code': 'Input_Token', 'unit_price': 0.05},
                {'dimension_code': 'Output_Token_Thinking', 'unit_price': 0.14}
            ]
        }
        result = extract_prices(spec)
        assert result['input_price'] == 0.05
        assert result['output_price'] == 0.14
    
    def test_extract_prices_with_zero_prices(self):
        """Test handling of zero prices"""
        from app.services.excel_exporter import extract_prices
        
        # Zero prices should be valid
        spec = {
            'prices': [
                {'dimension_code': 'input', 'unit_price': 0},
                {'dimension_code': 'output', 'unit_price': 0}
            ]
        }
        result = extract_prices(spec)
        assert result['input_price'] == 0
        assert result['output_price'] == 0
    
    def test_extract_prices_with_negative_prices(self):
        """Test handling of negative prices (should be accepted as-is)"""
        from app.services.excel_exporter import extract_prices
        
        # Negative prices should be accepted (might represent credits/discounts)
        spec = {
            'prices': [
                {'dimension_code': 'input', 'unit_price': -0.04}
            ]
        }
        result = extract_prices(spec)
        assert result['input_price'] == -0.04
    
    def test_extract_prices_with_large_prices(self):
        """Test handling of very large prices"""
        from app.services.excel_exporter import extract_prices
        
        # Large prices should be handled correctly
        spec = {
            'prices': [
                {'dimension_code': 'input', 'unit_price': 999999.99},
                {'dimension_code': 'output', 'unit_price': 1000000.00}
            ]
        }
        result = extract_prices(spec)
        assert result['input_price'] == 999999.99
        assert result['output_price'] == 1000000.00
    
    def test_extract_prices_with_decimal_precision(self):
        """Test handling of high decimal precision prices"""
        from app.services.excel_exporter import extract_prices
        
        # High precision prices should be preserved
        spec = {
            'prices': [
                {'dimension_code': 'input', 'unit_price': 0.0001234},
                {'dimension_code': 'output', 'unit_price': 0.0005678}
            ]
        }
        result = extract_prices(spec)
        assert abs(result['input_price'] - 0.0001234) < 1e-10
        assert abs(result['output_price'] - 0.0005678) < 1e-10
    
    def test_extract_prices_empty_spec(self):
        """Test with empty spec"""
        from app.services.excel_exporter import extract_prices
        
        # Empty spec should return all None
        spec = {}
        result = extract_prices(spec)
        assert result['input_price'] is None
        assert result['output_price'] is None
        assert result['non_token_price'] is None
        assert result['price_unit'] is None
        assert result['dimension_code'] is None
    
    def test_extract_prices_return_structure(self):
        """Test that return structure always has all required keys"""
        from app.services.excel_exporter import extract_prices
        
        # Test with various inputs
        test_specs = [
            {},
            {'prices': []},
            {'input_price': 0.04},
            {'prices': [{'dimension_code': 'input', 'unit_price': 0.04}]},
            {'prices': [{'dimension_code': 'character', 'unit_price': 0.001}]}
        ]
        
        for spec in test_specs:
            result = extract_prices(spec)
            assert 'input_price' in result
            assert 'output_price' in result
            assert 'non_token_price' in result
            assert 'price_unit' in result
            assert 'dimension_code' in result


class TestDiscountResolution:
    """Discount resolution tests - Requirements 5.1, 5.2"""
    
    def test_get_spec_discount_with_spec_level_discount(self):
        """Test that spec-level discount takes precedence (Requirement 5.1)"""
        from app.services.excel_exporter import get_spec_discount
        
        # Test with spec-level discount present
        spec_discounts = {
            'model1': {
                'spec1': 10.0,
                'spec2': 15.0
            },
            'model2': {
                'spec1': 20.0
            }
        }
        
        # Should return spec-level discount
        discount = get_spec_discount('model1', 'spec1', spec_discounts, 5.0)
        assert discount == 10.0
        
        discount = get_spec_discount('model1', 'spec2', spec_discounts, 5.0)
        assert discount == 15.0
        
        discount = get_spec_discount('model2', 'spec1', spec_discounts, 5.0)
        assert discount == 20.0
    
    def test_get_spec_discount_fallback_to_global(self):
        """Test fallback to global discount when spec-level not found (Requirement 5.2)"""
        from app.services.excel_exporter import get_spec_discount
        
        spec_discounts = {
            'model1': {
                'spec1': 10.0
            }
        }
        
        # Should return global discount when spec not found
        discount = get_spec_discount('model1', 'spec2', spec_discounts, 5.0)
        assert discount == 5.0
        
        # Should return global discount when model not found
        discount = get_spec_discount('model2', 'spec1', spec_discounts, 5.0)
        assert discount == 5.0
        
        # Should return global discount when both not found
        discount = get_spec_discount('model3', 'spec3', spec_discounts, 5.0)
        assert discount == 5.0
    
    def test_get_spec_discount_with_empty_spec_discounts(self):
        """Test with empty spec_discounts dict"""
        from app.services.excel_exporter import get_spec_discount
        
        # Empty dict should always return global discount
        discount = get_spec_discount('model1', 'spec1', {}, 5.0)
        assert discount == 5.0
        
        discount = get_spec_discount('any_model', 'any_spec', {}, 10.0)
        assert discount == 10.0
    
    def test_get_spec_discount_with_zero_discounts(self):
        """Test handling of zero discount values"""
        from app.services.excel_exporter import get_spec_discount
        
        spec_discounts = {
            'model1': {
                'spec1': 0.0
            }
        }
        
        # Zero spec-level discount should be returned (not treated as missing)
        discount = get_spec_discount('model1', 'spec1', spec_discounts, 5.0)
        assert discount == 0.0
        
        # Zero global discount should be returned when spec not found
        discount = get_spec_discount('model1', 'spec2', spec_discounts, 0.0)
        assert discount == 0.0
    
    def test_get_spec_discount_with_100_percent_discount(self):
        """Test handling of 100% discount (free)"""
        from app.services.excel_exporter import get_spec_discount
        
        spec_discounts = {
            'model1': {
                'spec1': 100.0
            }
        }
        
        # 100% discount should be handled correctly
        discount = get_spec_discount('model1', 'spec1', spec_discounts, 5.0)
        assert discount == 100.0
    
    def test_get_spec_discount_with_various_discount_values(self):
        """Test with various discount percentage values"""
        from app.services.excel_exporter import get_spec_discount
        
        spec_discounts = {
            'model1': {
                'spec1': 5.5,
                'spec2': 10.25,
                'spec3': 99.99
            }
        }
        
        # Test various discount values
        discount = get_spec_discount('model1', 'spec1', spec_discounts, 0.0)
        assert discount == 5.5
        
        discount = get_spec_discount('model1', 'spec2', spec_discounts, 0.0)
        assert discount == 10.25
        
        discount = get_spec_discount('model1', 'spec3', spec_discounts, 0.0)
        assert discount == 99.99
    
    def test_get_spec_discount_with_string_ids(self):
        """Test with string model_id and spec_id"""
        from app.services.excel_exporter import get_spec_discount
        
        spec_discounts = {
            'qwen-max': {
                'spec-001': 10.0,
                'spec-002': 15.0
            }
        }
        
        # String IDs should work correctly
        discount = get_spec_discount('qwen-max', 'spec-001', spec_discounts, 5.0)
        assert discount == 10.0
        
        discount = get_spec_discount('qwen-max', 'spec-002', spec_discounts, 5.0)
        assert discount == 15.0
    
    def test_get_spec_discount_with_numeric_ids(self):
        """Test with numeric model_id and spec_id (as strings)"""
        from app.services.excel_exporter import get_spec_discount
        
        spec_discounts = {
            '123': {
                '456': 10.0,
                '789': 15.0
            }
        }
        
        # Numeric IDs (as strings) should work
        discount = get_spec_discount('123', '456', spec_discounts, 5.0)
        assert discount == 10.0
        
        discount = get_spec_discount('123', '789', spec_discounts, 5.0)
        assert discount == 15.0
    
    def test_get_spec_discount_with_invalid_model_discounts_structure(self):
        """Test handling of invalid model_discounts structure"""
        from app.services.excel_exporter import get_spec_discount
        
        # Test with non-dict model_discounts value
        spec_discounts = {
            'model1': 'invalid'  # Should be a dict
        }
        
        # Should fall back to global discount
        discount = get_spec_discount('model1', 'spec1', spec_discounts, 5.0)
        assert discount == 5.0
        
        # Test with None value
        spec_discounts = {
            'model1': None
        }
        
        discount = get_spec_discount('model1', 'spec1', spec_discounts, 5.0)
        assert discount == 5.0
    
    def test_get_spec_discount_precedence(self):
        """Test that spec-level always takes precedence over global"""
        from app.services.excel_exporter import get_spec_discount
        
        spec_discounts = {
            'model1': {
                'spec1': 1.0,  # Very low spec discount
                'spec2': 99.0  # Very high spec discount
            }
        }
        
        # Low spec discount should override high global discount
        discount = get_spec_discount('model1', 'spec1', spec_discounts, 50.0)
        assert discount == 1.0
        
        # High spec discount should override low global discount
        discount = get_spec_discount('model1', 'spec2', spec_discounts, 0.0)
        assert discount == 99.0
    
    def test_get_spec_discount_examples_from_docstring(self):
        """Test examples from function docstring"""
        from app.services.excel_exporter import get_spec_discount
        
        # Example 1: get_spec_discount('model1', 'spec1', {'model1': {'spec1': 10.0}}, 5.0)
        discount = get_spec_discount('model1', 'spec1', {'model1': {'spec1': 10.0}}, 5.0)
        assert discount == 10.0
        
        # Example 2: get_spec_discount('model1', 'spec2', {'model1': {'spec1': 10.0}}, 5.0)
        discount = get_spec_discount('model1', 'spec2', {'model1': {'spec1': 10.0}}, 5.0)
        assert discount == 5.0
        
        # Example 3: get_spec_discount('model2', 'spec1', {}, 5.0)
        discount = get_spec_discount('model2', 'spec1', {}, 5.0)
        assert discount == 5.0
    
    def test_get_spec_discount_return_type(self):
        """Test that return type is always a float"""
        from app.services.excel_exporter import get_spec_discount
        
        spec_discounts = {
            'model1': {
                'spec1': 10.0
            }
        }
        
        # Test various scenarios
        discount = get_spec_discount('model1', 'spec1', spec_discounts, 5.0)
        assert isinstance(discount, float)
        
        discount = get_spec_discount('model1', 'spec2', spec_discounts, 5.0)
        assert isinstance(discount, float)
        
        discount = get_spec_discount('model2', 'spec1', {}, 5.0)
        assert isinstance(discount, float)


class TestHasAnyDiscount:
    """Tests for has_any_discount() function - Requirement 5.5"""
    
    def test_has_any_discount_no_discounts(self):
        """Test with no discounts at all (Requirement 5.5)"""
        from app.services.excel_exporter import has_any_discount
        
        # No global discount, no spec discounts
        result = has_any_discount({}, 0.0)
        assert result is False
        
        # Empty spec_discounts with zero global discount
        result = has_any_discount({}, 0)
        assert result is False
    
    def test_has_any_discount_global_discount_only(self):
        """Test with only global discount (Requirement 5.5)"""
        from app.services.excel_exporter import has_any_discount
        
        # Global discount > 0, no spec discounts
        result = has_any_discount({}, 5.0)
        assert result is True
        
        result = has_any_discount({}, 10.0)
        assert result is True
        
        result = has_any_discount({}, 100.0)
        assert result is True
        
        result = has_any_discount({}, 0.01)
        assert result is True
    
    def test_has_any_discount_spec_discount_only(self):
        """Test with only spec-level discounts (Requirement 5.5)"""
        from app.services.excel_exporter import has_any_discount
        
        # Spec discount > 0, no global discount
        spec_discounts = {'model1': {'spec1': 10.0}}
        result = has_any_discount(spec_discounts, 0.0)
        assert result is True
        
        # Multiple spec discounts
        spec_discounts = {
            'model1': {'spec1': 10.0, 'spec2': 15.0},
            'model2': {'spec1': 20.0}
        }
        result = has_any_discount(spec_discounts, 0.0)
        assert result is True
    
    def test_has_any_discount_both_discounts(self):
        """Test with both global and spec-level discounts (Requirement 5.5)"""
        from app.services.excel_exporter import has_any_discount
        
        # Both global and spec discounts present
        spec_discounts = {'model1': {'spec1': 10.0}}
        result = has_any_discount(spec_discounts, 5.0)
        assert result is True
    
    def test_has_any_discount_zero_spec_discounts(self):
        """Test with zero spec-level discounts (Requirement 5.5)"""
        from app.services.excel_exporter import has_any_discount
        
        # All spec discounts are zero, no global discount
        spec_discounts = {'model1': {'spec1': 0.0}}
        result = has_any_discount(spec_discounts, 0.0)
        assert result is False
        
        # Multiple zero spec discounts
        spec_discounts = {
            'model1': {'spec1': 0.0, 'spec2': 0.0},
            'model2': {'spec1': 0.0}
        }
        result = has_any_discount(spec_discounts, 0.0)
        assert result is False
    
    def test_has_any_discount_mixed_spec_discounts(self):
        """Test with mix of zero and non-zero spec discounts (Requirement 5.5)"""
        from app.services.excel_exporter import has_any_discount
        
        # Some zero, some non-zero spec discounts
        spec_discounts = {
            'model1': {'spec1': 0.0, 'spec2': 10.0},
            'model2': {'spec1': 0.0}
        }
        result = has_any_discount(spec_discounts, 0.0)
        assert result is True  # At least one spec discount > 0
    
    def test_has_any_discount_empty_model_discounts(self):
        """Test with empty model discount dicts"""
        from app.services.excel_exporter import has_any_discount
        
        # Model exists but has no specs
        spec_discounts = {'model1': {}}
        result = has_any_discount(spec_discounts, 0.0)
        assert result is False
        
        # Multiple models with empty specs
        spec_discounts = {'model1': {}, 'model2': {}}
        result = has_any_discount(spec_discounts, 0.0)
        assert result is False
    
    def test_has_any_discount_invalid_model_discounts_structure(self):
        """Test with invalid model_discounts structure"""
        from app.services.excel_exporter import has_any_discount
        
        # Non-dict model_discounts value
        spec_discounts = {'model1': 'invalid'}
        result = has_any_discount(spec_discounts, 0.0)
        assert result is False
        
        # List instead of dict
        spec_discounts = {'model1': [10.0, 20.0]}
        result = has_any_discount(spec_discounts, 0.0)
        assert result is False
        
        # None value
        spec_discounts = {'model1': None}
        result = has_any_discount(spec_discounts, 0.0)
        assert result is False
    
    def test_has_any_discount_negative_discounts(self):
        """Test handling of negative discount values"""
        from app.services.excel_exporter import has_any_discount
        
        # Negative global discount (should be treated as > 0)
        result = has_any_discount({}, -5.0)
        assert result is False  # Negative is not > 0
        
        # Negative spec discount
        spec_discounts = {'model1': {'spec1': -10.0}}
        result = has_any_discount(spec_discounts, 0.0)
        assert result is False  # Negative is not > 0
    
    def test_has_any_discount_very_small_discounts(self):
        """Test with very small discount values"""
        from app.services.excel_exporter import has_any_discount
        
        # Very small global discount
        result = has_any_discount({}, 0.0001)
        assert result is True
        
        # Very small spec discount
        spec_discounts = {'model1': {'spec1': 0.0001}}
        result = has_any_discount(spec_discounts, 0.0)
        assert result is True
    
    def test_has_any_discount_large_discounts(self):
        """Test with large discount values"""
        from app.services.excel_exporter import has_any_discount
        
        # 100% discount
        result = has_any_discount({}, 100.0)
        assert result is True
        
        # Over 100% discount (edge case)
        result = has_any_discount({}, 150.0)
        assert result is True
        
        spec_discounts = {'model1': {'spec1': 100.0}}
        result = has_any_discount(spec_discounts, 0.0)
        assert result is True
    
    def test_has_any_discount_return_type(self):
        """Test that return type is always a boolean"""
        from app.services.excel_exporter import has_any_discount
        
        # Test various scenarios
        result = has_any_discount({}, 0.0)
        assert isinstance(result, bool)
        
        result = has_any_discount({}, 5.0)
        assert isinstance(result, bool)
        
        result = has_any_discount({'model1': {'spec1': 10.0}}, 0.0)
        assert isinstance(result, bool)
        
        result = has_any_discount({'model1': {'spec1': 10.0}}, 5.0)
        assert isinstance(result, bool)
    
    def test_has_any_discount_examples_from_docstring(self):
        """Test examples from function docstring"""
        from app.services.excel_exporter import has_any_discount
        
        # Example 1: has_any_discount({}, 0.0)
        result = has_any_discount({}, 0.0)
        assert result is False
        
        # Example 2: has_any_discount({}, 5.0)
        result = has_any_discount({}, 5.0)
        assert result is True
        
        # Example 3: has_any_discount({'model1': {'spec1': 10.0}}, 0.0)
        result = has_any_discount({'model1': {'spec1': 10.0}}, 0.0)
        assert result is True
        
        # Example 4: has_any_discount({'model1': {'spec1': 0.0}}, 0.0)
        result = has_any_discount({'model1': {'spec1': 0.0}}, 0.0)
        assert result is False
    
    def test_has_any_discount_complex_nested_structure(self):
        """Test with complex nested discount structures"""
        from app.services.excel_exporter import has_any_discount
        
        # Large nested structure with one non-zero discount
        spec_discounts = {
            'model1': {'spec1': 0.0, 'spec2': 0.0, 'spec3': 0.0},
            'model2': {'spec1': 0.0, 'spec2': 0.0},
            'model3': {'spec1': 0.0, 'spec2': 5.0},  # One non-zero
            'model4': {'spec1': 0.0}
        }
        result = has_any_discount(spec_discounts, 0.0)
        assert result is True
        
        # Large nested structure with all zeros
        spec_discounts = {
            'model1': {'spec1': 0.0, 'spec2': 0.0, 'spec3': 0.0},
            'model2': {'spec1': 0.0, 'spec2': 0.0},
            'model3': {'spec1': 0.0, 'spec2': 0.0},
            'model4': {'spec1': 0.0}
        }
        result = has_any_discount(spec_discounts, 0.0)
        assert result is False


class TestPriceConversion:
    """Price unit conversion tests - Requirements 4.1, 4.2"""
    
    def test_convert_price_unit_thousand(self):
        """Test conversion with 'thousand' unit (Requirement 4.1)"""
        from app.services.excel_exporter import convert_price_unit
        
        # Test with 'thousand' - price should remain unchanged
        price, unit = convert_price_unit(0.5, 'thousand')
        assert price == 0.5
        assert unit == '千Token'
        
        # Test with different prices
        price, unit = convert_price_unit(0.04, 'thousand')
        assert price == 0.04
        assert unit == '千Token'
        
        price, unit = convert_price_unit(1.0, 'thousand')
        assert price == 1.0
        assert unit == '千Token'
        
        price, unit = convert_price_unit(100.5, 'thousand')
        assert price == 100.5
        assert unit == '千Token'
    
    def test_convert_price_unit_million(self):
        """Test conversion with 'million' unit (Requirement 4.2)"""
        from app.services.excel_exporter import convert_price_unit
        
        # Test with 'million' - price should be multiplied by 1000
        price, unit = convert_price_unit(0.5, 'million')
        assert price == 500.0
        assert unit == '百万Token'
        
        # Test with different prices
        price, unit = convert_price_unit(0.04, 'million')
        assert price == 40.0
        assert unit == '百万Token'
        
        price, unit = convert_price_unit(1.0, 'million')
        assert price == 1000.0
        assert unit == '百万Token'
        
        price, unit = convert_price_unit(0.001, 'million')
        assert price == 1.0
        assert unit == '百万Token'
    
    def test_convert_price_unit_none_price(self):
        """Test handling of None prices gracefully"""
        from app.services.excel_exporter import convert_price_unit
        
        # Test with None price and 'thousand'
        price, unit = convert_price_unit(None, 'thousand')
        assert price is None
        assert unit == '千Token'
        
        # Test with None price and 'million'
        price, unit = convert_price_unit(None, 'million')
        assert price is None
        assert unit == '百万Token'
    
    def test_convert_price_unit_zero_price(self):
        """Test handling of zero prices"""
        from app.services.excel_exporter import convert_price_unit
        
        # Zero should be handled correctly
        price, unit = convert_price_unit(0, 'thousand')
        assert price == 0
        assert unit == '千Token'
        
        price, unit = convert_price_unit(0, 'million')
        assert price == 0
        assert unit == '百万Token'
    
    def test_convert_price_unit_negative_price(self):
        """Test handling of negative prices"""
        from app.services.excel_exporter import convert_price_unit
        
        # Negative prices should be converted correctly
        price, unit = convert_price_unit(-0.5, 'thousand')
        assert price == -0.5
        assert unit == '千Token'
        
        price, unit = convert_price_unit(-0.5, 'million')
        assert price == -500.0
        assert unit == '百万Token'
    
    def test_convert_price_unit_large_price(self):
        """Test handling of very large prices"""
        from app.services.excel_exporter import convert_price_unit
        
        # Large prices should be handled correctly
        price, unit = convert_price_unit(999999.99, 'thousand')
        assert price == 999999.99
        assert unit == '千Token'
        
        price, unit = convert_price_unit(999999.99, 'million')
        assert price == 999999990.0
        assert unit == '百万Token'
    
    def test_convert_price_unit_small_price(self):
        """Test handling of very small prices"""
        from app.services.excel_exporter import convert_price_unit
        
        # Small prices should maintain precision
        price, unit = convert_price_unit(0.0001, 'thousand')
        assert price == 0.0001
        assert unit == '千Token'
        
        price, unit = convert_price_unit(0.0001, 'million')
        assert price == 0.1
        assert unit == '百万Token'
    
    def test_convert_price_unit_decimal_precision(self):
        """Test that decimal precision is maintained"""
        from app.services.excel_exporter import convert_price_unit
        
        # High precision prices should be preserved
        price, unit = convert_price_unit(0.0001234, 'thousand')
        assert abs(price - 0.0001234) < 1e-10
        assert unit == '千Token'
        
        price, unit = convert_price_unit(0.0001234, 'million')
        assert abs(price - 0.1234) < 1e-10
        assert unit == '百万Token'


class TestMonthlyUsageCalculation:
    """Monthly usage calculation tests - Requirement 6.2"""
    
    def test_calculate_monthly_usage_valid_integer(self):
        """Test calculation with valid integer string input"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Test with integer string
        result = calculate_monthly_usage("1000")
        assert result == "30000"
        
        result = calculate_monthly_usage("100")
        assert result == "3000"
        
        result = calculate_monthly_usage("1")
        assert result == "30"
    
    def test_calculate_monthly_usage_valid_float(self):
        """Test calculation with valid float string input"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Test with float string
        result = calculate_monthly_usage("500.5")
        assert result == "15015.0"
        
        result = calculate_monthly_usage("100.25")
        assert result == "3007.5"
        
        result = calculate_monthly_usage("0.5")
        assert result == "15.0"
    
    def test_calculate_monthly_usage_zero(self):
        """Test calculation with zero input"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Zero should return "0"
        result = calculate_monthly_usage("0")
        assert result == "0"
        
        result = calculate_monthly_usage("0.0")
        assert result == "0"
    
    def test_calculate_monthly_usage_empty_string(self):
        """Test handling of empty string input"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Empty string should return "-"
        result = calculate_monthly_usage("")
        assert result == "-"
    
    def test_calculate_monthly_usage_none(self):
        """Test handling of None input"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # None should return "-"
        result = calculate_monthly_usage(None)
        assert result == "-"
    
    def test_calculate_monthly_usage_whitespace(self):
        """Test handling of whitespace-only input"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Whitespace should return "-"
        result = calculate_monthly_usage("   ")
        assert result == "-"
        
        result = calculate_monthly_usage("\t")
        assert result == "-"
        
        result = calculate_monthly_usage("\n")
        assert result == "-"
    
    def test_calculate_monthly_usage_invalid_string(self):
        """Test handling of invalid string input"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Invalid strings should return "-"
        result = calculate_monthly_usage("invalid")
        assert result == "-"
        
        result = calculate_monthly_usage("abc")
        assert result == "-"
        
        result = calculate_monthly_usage("12.34.56")
        assert result == "-"
        
        result = calculate_monthly_usage("1000tokens")
        assert result == "-"
    
    def test_calculate_monthly_usage_negative_value(self):
        """Test calculation with negative values"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Negative values should be calculated correctly
        result = calculate_monthly_usage("-100")
        assert result == "-3000"
        
        result = calculate_monthly_usage("-50.5")
        assert result == "-1515.0"
    
    def test_calculate_monthly_usage_large_value(self):
        """Test calculation with very large values"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Large values should be handled correctly
        result = calculate_monthly_usage("1000000")
        assert result == "30000000"
        
        result = calculate_monthly_usage("999999.99")
        assert result == "29999999.7"
    
    def test_calculate_monthly_usage_small_value(self):
        """Test calculation with very small values"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Small values should maintain precision
        result = calculate_monthly_usage("0.001")
        assert result == "0.03"
        
        result = calculate_monthly_usage("0.0001")
        assert result == "0.003"
    
    def test_calculate_monthly_usage_scientific_notation(self):
        """Test calculation with scientific notation input"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Scientific notation should be parsed correctly
        result = calculate_monthly_usage("1e3")
        assert result == "30000.0"
        
        result = calculate_monthly_usage("1.5e2")
        assert result == "4500.0"
    
    def test_calculate_monthly_usage_with_leading_trailing_spaces(self):
        """Test calculation with leading/trailing spaces"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Leading/trailing spaces should be handled (strip happens in float conversion)
        result = calculate_monthly_usage("  1000  ")
        assert result == "30000"
        
        result = calculate_monthly_usage(" 500.5 ")
        assert result == "15015.0"
    
    def test_calculate_monthly_usage_formula_correctness(self):
        """Test that the formula (daily * 30) is correct"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Verify the multiplication by 30
        test_cases = [
            ("10", "300"),
            ("20", "600"),
            ("33.33", "999.9"),
            ("100", "3000"),
            ("1.5", "45.0")
        ]
        
        for daily, expected_monthly in test_cases:
            result = calculate_monthly_usage(daily)
            assert result == expected_monthly, f"Failed for daily={daily}: expected {expected_monthly}, got {result}"
    
    def test_calculate_monthly_usage_whole_number_formatting(self):
        """Test that whole numbers are formatted without decimal point"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Whole numbers should not have decimal point
        result = calculate_monthly_usage("100")
        assert result == "3000"
        assert "." not in result
        
        result = calculate_monthly_usage("10")
        assert result == "300"
        assert "." not in result
    
    def test_calculate_monthly_usage_decimal_formatting(self):
        """Test that decimal numbers are formatted with decimal point"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Decimal numbers should have decimal point
        result = calculate_monthly_usage("100.5")
        assert result == "3015.0"
        assert "." in result
        
        result = calculate_monthly_usage("10.1")
        assert result == "303.0"
        assert "." in result
    
    def test_calculate_monthly_usage_edge_cases(self):
        """Test various edge cases"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Test with special float values
        result = calculate_monthly_usage("0.0")
        assert result == "0"
        
        # Test with very precise decimals
        result = calculate_monthly_usage("1.123456789")
        # Should preserve precision in calculation
        expected = str(1.123456789 * 30)
        assert result == expected
    
    def test_calculate_monthly_usage_return_type(self):
        """Test that return type is always a string"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Test various inputs
        test_inputs = ["1000", "500.5", "0", "", None, "invalid"]
        
        for input_val in test_inputs:
            result = calculate_monthly_usage(input_val)
            assert isinstance(result, str), f"Return type should be str for input {input_val}"
    
    def test_calculate_monthly_usage_examples_from_docstring(self):
        """Test examples from function docstring"""
        from app.services.excel_exporter import calculate_monthly_usage
        
        # Example 1: calculate_monthly_usage("1000")
        result = calculate_monthly_usage("1000")
        assert result == "30000"
        
        # Example 2: calculate_monthly_usage("500.5")
        result = calculate_monthly_usage("500.5")
        assert result == "15015.0"
        
        # Example 3: calculate_monthly_usage("")
        result = calculate_monthly_usage("")
        assert result == "-"
        
        # Example 4: calculate_monthly_usage("invalid")
        result = calculate_monthly_usage("invalid")
        assert result == "-"
        
        # Example 5: calculate_monthly_usage("0")
        result = calculate_monthly_usage("0")
        assert result == "0"
    
    def test_convert_price_unit_return_type(self):
        """Test that return type is always a tuple"""
        from app.services.excel_exporter import convert_price_unit
        
        # Test various inputs
        result = convert_price_unit(0.5, 'thousand')
        assert isinstance(result, tuple)
        assert len(result) == 2
        
        result = convert_price_unit(0.5, 'million')
        assert isinstance(result, tuple)
        assert len(result) == 2
        
        result = convert_price_unit(None, 'thousand')
        assert isinstance(result, tuple)
        assert len(result) == 2
    
    def test_convert_price_unit_examples_from_docstring(self):
        """Test examples from function docstring"""
        from app.services.excel_exporter import convert_price_unit
        
        # Example 1: convert_price_unit(0.5, 'thousand')
        price, unit = convert_price_unit(0.5, 'thousand')
        assert price == 0.5
        assert unit == '千Token'
        
        # Example 2: convert_price_unit(0.5, 'million')
        price, unit = convert_price_unit(0.5, 'million')
        assert price == 500.0
        assert unit == '百万Token'
        
        # Example 3: convert_price_unit(None, 'thousand')
        price, unit = convert_price_unit(None, 'thousand')
        assert price is None
        assert unit == '千Token'



class TestMonthlyCalculations:
    """Monthly cost calculation tests - Requirements 6.3, 6.4, 6.5"""
    
    def test_calculate_monthly_cost_token_model_basic(self):
        """Test calculation for token-based models with basic inputs"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        # Test with both input and output prices, no discount
        spec = {
            'input_price': 0.04,
            'output_price': 0.12
        }
        result = calculate_monthly_cost(spec, "1000", 0.0, 'token')
        # (0.04 + 0.12) * 1000 * 30 * 1.0 = 4800.00
        assert result == "¥4,800.00"
        
        # Test with different prices
        spec = {
            'input_price': 0.02,
            'output_price': 0.08
        }
        result = calculate_monthly_cost(spec, "500", 0.0, 'token')
        # (0.02 + 0.08) * 500 * 30 * 1.0 = 1500.00
        assert result == "¥1,500.00"
    
    def test_calculate_monthly_cost_token_model_with_discount(self):
        """Test calculation for token-based models with discount applied"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        # Test with 10% discount
        spec = {
            'input_price': 0.04,
            'output_price': 0.12
        }
        result = calculate_monthly_cost(spec, "1000", 10.0, 'token')
        # (0.04 + 0.12) * 1000 * 30 * 0.9 = 4320.00
        assert result == "¥4,320.00"
        
        # Test with 50% discount
        result = calculate_monthly_cost(spec, "1000", 50.0, 'token')
        # (0.04 + 0.12) * 1000 * 30 * 0.5 = 2400.00
        assert result == "¥2,400.00"
        
        # Test with 100% discount (free)
        result = calculate_monthly_cost(spec, "1000", 100.0, 'token')
        # (0.04 + 0.12) * 1000 * 30 * 0.0 = 0.00
        assert result == "¥0.00"
    
    def test_calculate_monthly_cost_non_token_model_basic(self):
        """Test calculation for non-token models with basic inputs"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        # Test image generation (price per image)
        spec = {
            'non_token_price': 0.08
        }
        result = calculate_monthly_cost(spec, "500", 0.0, 'image')
        # 0.08 * 500 * 30 * 1.0 = 1200.00
        assert result == "¥1,200.00"
        
        # Test TTS (price per character)
        spec = {
            'non_token_price': 0.001
        }
        result = calculate_monthly_cost(spec, "10000", 0.0, 'character')
        # 0.001 * 10000 * 30 * 1.0 = 300.00
        assert result == "¥300.00"
        
        # Test ASR (price per second)
        spec = {
            'non_token_price': 0.002
        }
        result = calculate_monthly_cost(spec, "3600", 0.0, 'audio')
        # 0.002 * 3600 * 30 * 1.0 = 216.00
        assert result == "¥216.00"
        
        # Test video generation (price per second)
        spec = {
            'non_token_price': 0.5
        }
        result = calculate_monthly_cost(spec, "100", 0.0, 'video')
        # 0.5 * 100 * 30 * 1.0 = 1500.00
        assert result == "¥1,500.00"
    
    def test_calculate_monthly_cost_non_token_model_with_discount(self):
        """Test calculation for non-token models with discount applied"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        # Test with 20% discount
        spec = {
            'non_token_price': 0.08
        }
        result = calculate_monthly_cost(spec, "500", 20.0, 'image')
        # 0.08 * 500 * 30 * 0.8 = 960.00
        assert result == "¥960.00"
        
        # Test with 15% discount
        spec = {
            'non_token_price': 0.001
        }
        result = calculate_monthly_cost(spec, "10000", 15.0, 'character')
        # 0.001 * 10000 * 30 * 0.85 = 255.00
        assert result == "¥255.00"
    
    def test_calculate_monthly_cost_empty_daily_usage(self):
        """Test handling of empty daily usage"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        spec = {
            'input_price': 0.04,
            'output_price': 0.12
        }
        
        # Empty string should return "-"
        result = calculate_monthly_cost(spec, "", 0.0, 'token')
        assert result == "-"
        
        # None should return "-"
        result = calculate_monthly_cost(spec, None, 0.0, 'token')
        assert result == "-"
        
        # Whitespace should return "-"
        result = calculate_monthly_cost(spec, "   ", 0.0, 'token')
        assert result == "-"
    
    def test_calculate_monthly_cost_invalid_daily_usage(self):
        """Test handling of invalid daily usage values"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        spec = {
            'input_price': 0.04,
            'output_price': 0.12
        }
        
        # Invalid string should return "-"
        result = calculate_monthly_cost(spec, "invalid", 0.0, 'token')
        assert result == "-"
        
        result = calculate_monthly_cost(spec, "abc123", 0.0, 'token')
        assert result == "-"
        
        result = calculate_monthly_cost(spec, "12.34.56", 0.0, 'token')
        assert result == "-"
    
    def test_calculate_monthly_cost_zero_daily_usage(self):
        """Test handling of zero daily usage"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        spec = {
            'input_price': 0.04,
            'output_price': 0.12
        }
        
        # Zero should return "-" (no usage means no cost to display)
        result = calculate_monthly_cost(spec, "0", 0.0, 'token')
        assert result == "-"
        
        result = calculate_monthly_cost(spec, "0.0", 0.0, 'token')
        assert result == "-"
    
    def test_calculate_monthly_cost_negative_daily_usage(self):
        """Test handling of negative daily usage"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        spec = {
            'input_price': 0.04,
            'output_price': 0.12
        }
        
        # Negative values should return "-"
        result = calculate_monthly_cost(spec, "-100", 0.0, 'token')
        assert result == "-"
        
        result = calculate_monthly_cost(spec, "-50.5", 0.0, 'token')
        assert result == "-"
    
    def test_calculate_monthly_cost_missing_prices_token_model(self):
        """Test handling of missing prices for token models"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        # Both prices missing should return "-"
        spec = {}
        result = calculate_monthly_cost(spec, "1000", 0.0, 'token')
        assert result == "-"
        
        # Only input price present (output is None) - should calculate with output as 0
        spec = {
            'input_price': 0.04
        }
        result = calculate_monthly_cost(spec, "1000", 0.0, 'token')
        # 0.04 * 1000 * 30 * 1.0 = 1200.00
        assert result == "¥1,200.00"
        
        # Only output price present (input is None) - should calculate with input as 0
        spec = {
            'output_price': 0.12
        }
        result = calculate_monthly_cost(spec, "1000", 0.0, 'token')
        # 0.12 * 1000 * 30 * 1.0 = 3600.00
        assert result == "¥3,600.00"
    
    def test_calculate_monthly_cost_missing_prices_non_token_model(self):
        """Test handling of missing prices for non-token models"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        # Missing non_token_price should return "-"
        spec = {}
        result = calculate_monthly_cost(spec, "500", 0.0, 'image')
        assert result == "-"
        
        spec = {
            'input_price': 0.04,  # Wrong price type for non-token model
            'output_price': 0.12
        }
        result = calculate_monthly_cost(spec, "500", 0.0, 'image')
        assert result == "-"
    
    def test_calculate_monthly_cost_float_daily_usage(self):
        """Test calculation with float daily usage values"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        spec = {
            'input_price': 0.04,
            'output_price': 0.12
        }
        
        # Test with decimal daily usage
        result = calculate_monthly_cost(spec, "1000.5", 0.0, 'token')
        # (0.04 + 0.12) * 1000.5 * 30 * 1.0 = 4802.40
        assert result == "¥4,802.40"
        
        result = calculate_monthly_cost(spec, "500.25", 10.0, 'token')
        # (0.04 + 0.12) * 500.25 * 30 * 0.9 = 2161.08
        assert result == "¥2,161.08"  # Rounded to 2 decimal places
    
    def test_calculate_monthly_cost_large_values(self):
        """Test calculation with very large values"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        spec = {
            'input_price': 0.04,
            'output_price': 0.12
        }
        
        # Test with large daily usage
        result = calculate_monthly_cost(spec, "1000000", 0.0, 'token')
        # (0.04 + 0.12) * 1000000 * 30 * 1.0 = 4800000.00
        assert result == "¥4,800,000.00"
        
        # Verify thousands separator is present
        assert "," in result
    
    def test_calculate_monthly_cost_small_values(self):
        """Test calculation with very small values"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        spec = {
            'input_price': 0.0001,
            'output_price': 0.0002
        }
        
        # Test with small prices and usage
        result = calculate_monthly_cost(spec, "10", 0.0, 'token')
        # (0.0001 + 0.0002) * 10 * 30 * 1.0 = 0.09
        assert result == "¥0.09"
        
        # Test very small result
        spec = {
            'non_token_price': 0.00001
        }
        result = calculate_monthly_cost(spec, "1", 0.0, 'image')
        # 0.00001 * 1 * 30 * 1.0 = 0.0003
        assert result == "¥0.00"  # Rounds to 0.00
    
    def test_calculate_monthly_cost_discount_rate_calculation(self):
        """Test that discount rate is calculated correctly"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        spec = {
            'input_price': 0.10,
            'output_price': 0.10
        }
        
        # Test various discount percentages
        # 0% discount: rate = 1.0
        result = calculate_monthly_cost(spec, "100", 0.0, 'token')
        # (0.10 + 0.10) * 100 * 30 * 1.0 = 600.00
        assert result == "¥600.00"
        
        # 25% discount: rate = 0.75
        result = calculate_monthly_cost(spec, "100", 25.0, 'token')
        # (0.10 + 0.10) * 100 * 30 * 0.75 = 450.00
        assert result == "¥450.00"
        
        # 75% discount: rate = 0.25
        result = calculate_monthly_cost(spec, "100", 75.0, 'token')
        # (0.10 + 0.10) * 100 * 30 * 0.25 = 150.00
        assert result == "¥150.00"
    
    def test_calculate_monthly_cost_formula_token_model(self):
        """Test that the formula for token models is correct: (input_price + output_price) × daily × 30 × discount_rate"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        # Test case 1
        spec = {
            'input_price': 0.05,
            'output_price': 0.15
        }
        daily = 200
        discount = 10.0
        discount_rate = (100 - discount) / 100
        expected_cost = (0.05 + 0.15) * daily * 30 * discount_rate
        
        result = calculate_monthly_cost(spec, str(daily), discount, 'token')
        assert result == f"¥{expected_cost:,.2f}"
        
        # Test case 2
        spec = {
            'input_price': 0.03,
            'output_price': 0.09
        }
        daily = 1500
        discount = 20.0
        discount_rate = (100 - discount) / 100
        expected_cost = (0.03 + 0.09) * daily * 30 * discount_rate
        
        result = calculate_monthly_cost(spec, str(daily), discount, 'token')
        assert result == f"¥{expected_cost:,.2f}"
    
    def test_calculate_monthly_cost_formula_non_token_model(self):
        """Test that the formula for non-token models is correct: non_token_price × daily × 30 × discount_rate"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        # Test case 1
        spec = {
            'non_token_price': 0.10
        }
        daily = 300
        discount = 15.0
        discount_rate = (100 - discount) / 100
        expected_cost = 0.10 * daily * 30 * discount_rate
        
        result = calculate_monthly_cost(spec, str(daily), discount, 'image')
        assert result == f"¥{expected_cost:,.2f}"
        
        # Test case 2
        spec = {
            'non_token_price': 0.005
        }
        daily = 5000
        discount = 5.0
        discount_rate = (100 - discount) / 100
        expected_cost = 0.005 * daily * 30 * discount_rate
        
        result = calculate_monthly_cost(spec, str(daily), discount, 'character')
        assert result == f"¥{expected_cost:,.2f}"
    
    def test_calculate_monthly_cost_currency_formatting(self):
        """Test that currency is formatted correctly with ¥ symbol, thousands separator, and 2 decimal places"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        spec = {
            'input_price': 0.04,
            'output_price': 0.12
        }
        
        # Test formatting
        result = calculate_monthly_cost(spec, "1000", 0.0, 'token')
        
        # Should start with ¥
        assert result.startswith("¥")
        
        # Should have thousands separator for large numbers
        result_large = calculate_monthly_cost(spec, "10000", 0.0, 'token')
        assert "," in result_large
        
        # Should have exactly 2 decimal places
        assert result.endswith(".00") or result[-3] == "."
        
        # Test various amounts
        test_cases = [
            ("100", "¥480.00"),
            ("1000", "¥4,800.00"),
            ("10000", "¥48,000.00"),
        ]
        
        for daily, expected in test_cases:
            result = calculate_monthly_cost(spec, daily, 0.0, 'token')
            assert result == expected
    
    def test_calculate_monthly_cost_return_type(self):
        """Test that return type is always a string"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        spec = {
            'input_price': 0.04,
            'output_price': 0.12
        }
        
        # Test various inputs
        test_inputs = [
            ("1000", 0.0, 'token'),
            ("", 0.0, 'token'),
            ("invalid", 0.0, 'token'),
            ("500", 10.0, 'token'),
        ]
        
        for daily, discount, price_type in test_inputs:
            result = calculate_monthly_cost(spec, daily, discount, price_type)
            assert isinstance(result, str), f"Return type should be str for input {daily}"
    
    def test_calculate_monthly_cost_all_price_types(self):
        """Test calculation works for all price types"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        # Token-based price types
        token_spec = {
            'input_price': 0.04,
            'output_price': 0.12
        }
        
        for price_type in ['token']:
            result = calculate_monthly_cost(token_spec, "1000", 0.0, price_type)
            assert result.startswith("¥")
            assert result != "-"
        
        # Non-token price types
        non_token_spec = {
            'non_token_price': 0.08
        }
        
        for price_type in ['image', 'character', 'audio', 'video']:
            result = calculate_monthly_cost(non_token_spec, "500", 0.0, price_type)
            assert result.startswith("¥")
            assert result != "-"
    
    def test_calculate_monthly_cost_edge_case_zero_prices(self):
        """Test handling of zero prices"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        # Zero prices should still calculate (result will be ¥0.00)
        spec = {
            'input_price': 0.0,
            'output_price': 0.0
        }
        result = calculate_monthly_cost(spec, "1000", 0.0, 'token')
        assert result == "¥0.00"
        
        spec = {
            'non_token_price': 0.0
        }
        result = calculate_monthly_cost(spec, "500", 0.0, 'image')
        assert result == "¥0.00"
    
    def test_calculate_monthly_cost_scientific_notation(self):
        """Test calculation with scientific notation in daily usage"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        spec = {
            'input_price': 0.04,
            'output_price': 0.12
        }
        
        # Scientific notation should be parsed correctly
        result = calculate_monthly_cost(spec, "1e3", 0.0, 'token')
        # 1e3 = 1000, so (0.04 + 0.12) * 1000 * 30 * 1.0 = 4800.00
        assert result == "¥4,800.00"
        
        result = calculate_monthly_cost(spec, "1.5e2", 0.0, 'token')
        # 1.5e2 = 150, so (0.04 + 0.12) * 150 * 30 * 1.0 = 720.00
        assert result == "¥720.00"
    
    def test_calculate_monthly_cost_with_leading_trailing_spaces(self):
        """Test calculation with leading/trailing spaces in daily usage"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        spec = {
            'input_price': 0.04,
            'output_price': 0.12
        }
        
        # Leading/trailing spaces should be handled
        result = calculate_monthly_cost(spec, "  1000  ", 0.0, 'token')
        assert result == "¥4,800.00"
        
        result = calculate_monthly_cost(spec, " 500 ", 10.0, 'token')
        # (0.04 + 0.12) * 500 * 30 * 0.9 = 2160.00
        assert result == "¥2,160.00"
    
    def test_calculate_monthly_cost_examples_from_docstring(self):
        """Test examples from function docstring"""
        from app.services.excel_exporter import calculate_monthly_cost
        
        # Example 1: Token model with discount
        spec = {'input_price': 0.04, 'output_price': 0.12}
        result = calculate_monthly_cost(spec, "1000", 10.0, 'token')
        assert result == "¥4,320.00"
        
        # Example 2: Non-token model without discount
        spec = {'non_token_price': 0.08}
        result = calculate_monthly_cost(spec, "500", 0.0, 'image')
        assert result == "¥1,200.00"
        
        # Example 3: Missing prices
        result = calculate_monthly_cost({}, "1000", 0.0, 'token')
        assert result == "-"
        
        # Example 4: Empty daily usage
        spec = {'input_price': 0.04, 'output_price': 0.12}
        result = calculate_monthly_cost(spec, "", 0.0, 'token')
        assert result == "-"
    
    def test_calculate_monthly_cost_integration_with_extract_prices(self):
        """Test that calculate_monthly_cost works with output from extract_prices"""
        from app.services.excel_exporter import extract_prices, calculate_monthly_cost
        
        # Test with new prices array format
        spec_data = {
            'prices': [
                {'dimension_code': 'input', 'unit_price': 0.04},
                {'dimension_code': 'output', 'unit_price': 0.12}
            ]
        }
        
        extracted = extract_prices(spec_data)
        result = calculate_monthly_cost(extracted, "1000", 10.0, 'token')
        assert result == "¥4,320.00"
        
        # Test with non-token prices
        spec_data = {
            'prices': [
                {'dimension_code': 'image_count', 'unit_price': 0.08}
            ]
        }
        
        extracted = extract_prices(spec_data)
        result = calculate_monthly_cost(extracted, "500", 0.0, 'image')
        assert result == "¥1,200.00"



class TestRenderCategoryHeader:
    """Tests for render_category_header() function - Requirements 2.1, 2.2, 2.3"""
    
    def test_render_category_header_basic(self):
        """Test basic category header rendering with valid inputs"""
        from app.services.excel_exporter import render_category_header
        from openpyxl import Workbook
        
        # Create a test workbook
        wb = Workbook()
        ws = wb.active
        
        # Render header for text_qwen category with 5 items at row 1, spanning 11 columns
        next_row = render_category_header(ws, 'text_qwen', 5, 1, 11)
        
        # Verify next row is incremented
        assert next_row == 2
        
        # Verify cell value contains icon, name, and count
        cell_value = ws['A1'].value
        assert '💬' in cell_value  # Icon
        assert '文本生成-通义千问' in cell_value  # Name
        assert '共5项' in cell_value  # Count
        assert cell_value == '💬 文本生成-通义千问 (共5项)'
        
        # Verify cells are merged (A1:K1 for 11 columns)
        assert 'A1:K1' in ws.merged_cells
    
    def test_render_category_header_styling(self):
        """Test that correct styling is applied to category header (Requirement 2.3)"""
        from app.services.excel_exporter import render_category_header
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Render header
        render_category_header(ws, 'image_gen', 3, 1, 8)
        
        cell = ws['A1']
        
        # Verify font styling: 12pt, bold
        assert cell.font.size == 12
        assert cell.font.bold is True
        assert cell.font.name == '微软雅黑'
        
        # Verify background color: #E7E6E6
        # openpyxl may use different alpha channel prefixes (00, FF, or none)
        assert cell.fill.start_color.rgb in ['FFE7E6E6', '00E7E6E6', 'E7E6E6']
        assert cell.fill.fill_type == 'solid'
        
        # Verify alignment: left horizontal, center vertical
        assert cell.alignment.horizontal == 'left'
        assert cell.alignment.vertical == 'center'
    
    def test_render_category_header_all_categories(self):
        """Test rendering headers for all 12 categories"""
        from app.services.excel_exporter import render_category_header, CATEGORY_CONFIG
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        current_row = 1
        
        # Test all categories
        for category_key, config in CATEGORY_CONFIG.items():
            next_row = render_category_header(ws, category_key, 10, current_row, 11)
            
            # Verify next row is incremented
            assert next_row == current_row + 1
            
            # Verify cell value contains correct icon and name
            cell_value = ws[f'A{current_row}'].value
            assert config['icon'] in cell_value
            assert config['name'] in cell_value
            assert '共10项' in cell_value
            
            current_row = next_row + 1  # Add spacing between headers
    
    def test_render_category_header_different_item_counts(self):
        """Test rendering with different item counts"""
        from app.services.excel_exporter import render_category_header
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Test with 0 items
        render_category_header(ws, 'text_qwen', 0, 1, 11)
        assert '共0项' in ws['A1'].value
        
        # Test with 1 item
        render_category_header(ws, 'image_gen', 1, 2, 11)
        assert '共1项' in ws['A2'].value
        
        # Test with large number
        render_category_header(ws, 'tts', 999, 3, 11)
        assert '共999项' in ws['A3'].value
    
    def test_render_category_header_different_column_counts(self):
        """Test rendering with different column counts"""
        from app.services.excel_exporter import render_category_header
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Test with 5 columns (A1:E1)
        render_category_header(ws, 'text_qwen', 5, 1, 5)
        assert 'A1:E1' in ws.merged_cells
        
        # Test with 13 columns (A2:M2)
        render_category_header(ws, 'image_gen', 3, 2, 13)
        assert 'A2:M2' in ws.merged_cells
        
        # Test with 1 column (no merge needed, but should still work)
        render_category_header(ws, 'tts', 2, 3, 1)
        # Single cell merge creates a range like A3:A3
        assert 'A3:A3' in ws.merged_cells
    
    def test_render_category_header_different_row_positions(self):
        """Test rendering at different row positions"""
        from app.services.excel_exporter import render_category_header
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Test at row 1
        next_row = render_category_header(ws, 'text_qwen', 5, 1, 11)
        assert next_row == 2
        assert ws['A1'].value is not None
        
        # Test at row 10
        next_row = render_category_header(ws, 'image_gen', 3, 10, 11)
        assert next_row == 11
        assert ws['A10'].value is not None
        
        # Test at row 100
        next_row = render_category_header(ws, 'tts', 7, 100, 11)
        assert next_row == 101
        assert ws['A100'].value is not None
    
    def test_render_category_header_invalid_category_key(self):
        """Test handling of invalid category key"""
        from app.services.excel_exporter import render_category_header
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Should handle gracefully with a default icon and use the key as name
        next_row = render_category_header(ws, 'invalid_category', 5, 1, 11)
        
        # Should still return next row
        assert next_row == 2
        
        # Should have some value (fallback behavior)
        cell_value = ws['A1'].value
        assert cell_value is not None
        assert '共5项' in cell_value
        # Should use default icon and the category key as name
        assert '📋' in cell_value
        assert 'invalid_category' in cell_value
    
    def test_render_category_header_specific_categories(self):
        """Test specific category examples from requirements"""
        from app.services.excel_exporter import render_category_header
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Test text_qwen (Requirement 2.4)
        render_category_header(ws, 'text_qwen', 5, 1, 11)
        assert ws['A1'].value == '💬 文本生成-通义千问 (共5项)'
        
        # Test image_gen
        render_category_header(ws, 'image_gen', 3, 2, 11)
        assert ws['A2'].value == '🎨 图像生成 (共3项)'
        
        # Test tts
        render_category_header(ws, 'tts', 2, 3, 11)
        assert ws['A3'].value == '🔊 语音合成 (共2项)'
        
        # Test asr
        render_category_header(ws, 'asr', 1, 4, 11)
        assert ws['A4'].value == '🎤 语音识别与翻译 (共1项)'
        
        # Test video_gen
        render_category_header(ws, 'video_gen', 4, 5, 11)
        assert ws['A5'].value == '🎬 视频生成 (共4项)'
    
    def test_render_category_header_return_value(self):
        """Test that function returns correct next row number"""
        from app.services.excel_exporter import render_category_header
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Test sequential rendering
        row = 1
        row = render_category_header(ws, 'text_qwen', 5, row, 11)
        assert row == 2
        
        row = render_category_header(ws, 'image_gen', 3, row, 11)
        assert row == 3
        
        row = render_category_header(ws, 'tts', 2, row, 11)
        assert row == 4
        
        # Verify all three headers were rendered
        assert ws['A1'].value is not None
        assert ws['A2'].value is not None
        assert ws['A3'].value is not None
    
    def test_render_category_header_preserves_other_cells(self):
        """Test that rendering doesn't affect other cells"""
        from app.services.excel_exporter import render_category_header
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Set some values in other cells
        ws['A5'].value = 'Test Value'
        ws['B1'].value = 'Another Value'
        
        # Render header at row 1
        render_category_header(ws, 'text_qwen', 5, 1, 11)
        
        # Verify other cells are preserved
        assert ws['A5'].value == 'Test Value'
        # Note: B1 will be part of the merged cell A1:K1, so it might be affected
        # This is expected behavior for merged cells



class TestRenderTokenBasedTable:
    """Tests for render_token_based_table() function - Requirements 1.1, 5.3, 9.2, 9.3"""
    
    def test_render_token_table_without_discounts(self):
        """Test rendering token-based table without any discounts"""
        from app.services.excel_exporter import render_token_based_table
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data - single model with one spec
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-max',
                    'name': 'qwen-max'
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ],
                        'remark': '测试备注'
                    }
                ]
            }
        ]
        
        # No discounts
        spec_discounts = {}
        daily_usages = {}
        global_discount = 0.0
        
        # Render table
        next_row = render_token_based_table(
            ws, models, 1, 'thousand', spec_discounts, daily_usages, global_discount
        )
        
        # Verify headers (without discount columns)
        expected_headers = [
            '序号', '模型名称', '模式', 'Token范围',
            '输入单价', '输出单价', 
            '日估计用量', '预估月用量', '预估月费', '备注'
        ]
        
        for col_idx, expected_header in enumerate(expected_headers, 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            assert cell_value == expected_header, f"Column {col_idx} header mismatch: expected '{expected_header}', got '{cell_value}'"
        
        # Verify header styling
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert header_cell.font.color.rgb in ['FFFFFFFF', '00FFFFFF', 'FFFFFF']
        assert header_cell.fill.start_color.rgb in ['FF4472C4', '004472C4', '4472C4']
        assert header_cell.alignment.horizontal == 'center'
        
        # Verify data row
        assert ws.cell(row=2, column=1).value == 1  # 序号
        assert ws.cell(row=2, column=2).value == 'qwen-max'  # 模型名称
        assert ws.cell(row=2, column=3).value == '标准模式'  # 模式
        assert ws.cell(row=2, column=4).value == '0-128K'  # Token范围
        assert '¥0.0400/千Token' in str(ws.cell(row=2, column=5).value)  # 输入单价
        assert '¥0.1200/千Token' in str(ws.cell(row=2, column=6).value)  # 输出单价
        assert ws.cell(row=2, column=7).value == '-'  # 日估计用量 (not provided)
        assert ws.cell(row=2, column=8).value == '-'  # 预估月用量
        assert ws.cell(row=2, column=9).value == '-'  # 预估月费
        assert ws.cell(row=2, column=10).value == '测试备注'  # 备注
        
        # Verify next row is correct
        assert next_row == 3
    
    def test_render_token_table_with_discounts(self):
        """Test rendering token-based table with discounts (Requirement 5.3)"""
        from app.services.excel_exporter import render_token_based_table
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-plus'
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.02},
                            {'dimension_code': 'output', 'unit_price': 0.06}
                        ],
                        'remark': ''
                    }
                ]
            }
        ]
        
        # With global discount
        spec_discounts = {}
        daily_usages = {}
        global_discount = 10.0  # 10% discount
        
        # Render table
        next_row = render_token_based_table(
            ws, models, 1, 'thousand', spec_discounts, daily_usages, global_discount
        )
        
        # Verify headers (with discount columns)
        expected_headers = [
            '序号', '模型名称', '模式', 'Token范围', 
            '输入单价', '输出单价', '折扣', '折后输入', '折后输出',
            '日估计用量', '预估月用量', '预估月费', '备注'
        ]
        
        for col_idx, expected_header in enumerate(expected_headers, 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            assert cell_value == expected_header
        
        # Verify discount columns in data row
        assert ws.cell(row=2, column=7).value == '10.0%'  # 折扣
        # 折后输入: 0.02 * 0.9 = 0.018
        assert '¥0.0180/千Token' in str(ws.cell(row=2, column=8).value)
        # 折后输出: 0.06 * 0.9 = 0.054
        assert '¥0.0540/千Token' in str(ws.cell(row=2, column=9).value)
    
    def test_render_token_table_with_spec_level_discount(self):
        """Test that spec-level discount takes precedence over global discount"""
        from app.services.excel_exporter import render_token_based_table
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-max'
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ]
                    }
                ]
            }
        ]
        
        # Spec-level discount should override global
        spec_discounts = {
            'model1': {
                'spec1': 20.0  # 20% spec-level discount
            }
        }
        daily_usages = {}
        global_discount = 10.0  # 10% global discount
        
        next_row = render_token_based_table(
            ws, models, 1, 'thousand', spec_discounts, daily_usages, global_discount
        )
        
        # Verify spec-level discount is used (20%, not 10%)
        assert ws.cell(row=2, column=7).value == '20.0%'
        # 折后输入: 0.04 * 0.8 = 0.032
        assert '¥0.0320/千Token' in str(ws.cell(row=2, column=8).value)
    
    def test_render_token_table_with_daily_usage(self):
        """Test rendering with daily usage and monthly calculations"""
        from app.services.excel_exporter import render_token_based_table
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-max'
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {
            'model1': {
                'spec1': '1000'  # 1000 tokens per day
            }
        }
        global_discount = 0.0
        
        next_row = render_token_based_table(
            ws, models, 1, 'thousand', spec_discounts, daily_usages, global_discount
        )
        
        # Verify daily usage
        assert ws.cell(row=2, column=7).value == '1000'
        
        # Verify monthly usage (1000 * 30 = 30000)
        assert ws.cell(row=2, column=8).value == '30000'
        
        # Verify monthly cost: (0.04 + 0.12) * 1000 * 30 = 4800
        monthly_cost = ws.cell(row=2, column=9).value
        assert '¥4,800.00' in monthly_cost
    
    def test_render_token_table_with_million_unit(self):
        """Test price unit conversion to million tokens (Requirement 4.2)"""
        from app.services.excel_exporter import render_token_based_table
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-max'
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {}
        global_discount = 0.0
        
        # Use 'million' price unit
        next_row = render_token_based_table(
            ws, models, 1, 'million', spec_discounts, daily_usages, global_discount
        )
        
        # Verify prices are multiplied by 1000 and unit is 百万Token
        # 0.04 * 1000 = 40.0
        assert '¥40.0000/百万Token' in str(ws.cell(row=2, column=5).value)
        # 0.12 * 1000 = 120.0
        assert '¥120.0000/百万Token' in str(ws.cell(row=2, column=6).value)
    
    def test_render_token_table_multiple_models(self):
        """Test rendering table with multiple models and specs"""
        from app.services.excel_exporter import render_token_based_table
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-max'
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ]
                    },
                    {
                        'id': 'spec2',
                        'mode': '思考模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output_token_thinking', 'unit_price': 0.24}
                        ]
                    }
                ]
            },
            {
                'model': {
                    'id': 'model2',
                    'model_name': 'qwen-plus'
                },
                'specs': [
                    {
                        'id': 'spec3',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.02},
                            {'dimension_code': 'output', 'unit_price': 0.06}
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {}
        global_discount = 0.0
        
        next_row = render_token_based_table(
            ws, models, 1, 'thousand', spec_discounts, daily_usages, global_discount
        )
        
        # Verify we have 3 data rows (2 specs from model1 + 1 spec from model2)
        assert next_row == 5  # 1 header + 3 data rows + 1
        
        # Verify row numbers
        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=3, column=1).value == 2
        assert ws.cell(row=4, column=1).value == 3
        
        # Verify model names
        assert ws.cell(row=2, column=2).value == 'qwen-max'
        assert ws.cell(row=3, column=2).value == 'qwen-max'
        assert ws.cell(row=4, column=2).value == 'qwen-plus'
        
        # Verify modes
        assert ws.cell(row=2, column=3).value == '标准模式'
        assert ws.cell(row=3, column=3).value == '思考模式'
        assert ws.cell(row=4, column=3).value == '标准模式'
    
    def test_render_token_table_with_legacy_token_range(self):
        """Test handling of legacy token_range field"""
        from app.services.excel_exporter import render_token_based_table
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-max'
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_range': '0-32K',  # Legacy field
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {}
        global_discount = 0.0
        
        next_row = render_token_based_table(
            ws, models, 1, 'thousand', spec_discounts, daily_usages, global_discount
        )
        
        # Verify legacy token_range is used
        assert ws.cell(row=2, column=4).value == '0-32K'
    
    def test_render_token_table_with_missing_prices(self):
        """Test handling of missing price data"""
        from app.services.excel_exporter import render_token_based_table
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-max'
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': []  # No prices
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {}
        global_discount = 0.0
        
        next_row = render_token_based_table(
            ws, models, 1, 'thousand', spec_discounts, daily_usages, global_discount
        )
        
        # Verify missing prices show as '-'
        assert ws.cell(row=2, column=5).value == '-'  # 输入单价
        assert ws.cell(row=2, column=6).value == '-'  # 输出单价
    
    def test_render_token_table_data_cell_styling(self):
        """Test that data cells have correct styling (Requirement 9.3)"""
        from app.services.excel_exporter import render_token_based_table
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-max'
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {}
        global_discount = 0.0
        
        next_row = render_token_based_table(
            ws, models, 1, 'thousand', spec_discounts, daily_usages, global_discount
        )
        
        # Check data cell styling
        data_cell = ws.cell(row=2, column=2)  # 模型名称
        assert data_cell.font.size == 10
        assert data_cell.font.name == '微软雅黑'
        assert data_cell.alignment.horizontal == 'left'
        assert data_cell.border.left.style == 'thin'
        
        # Check number cell alignment
        number_cell = ws.cell(row=2, column=1)  # 序号
        assert number_cell.alignment.horizontal == 'center'
        
        price_cell = ws.cell(row=2, column=5)  # 输入单价
        assert price_cell.alignment.horizontal == 'right'
    
    def test_render_token_table_empty_models_list(self):
        """Test rendering with empty models list"""
        from app.services.excel_exporter import render_token_based_table
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        models = []
        spec_discounts = {}
        daily_usages = {}
        global_discount = 0.0
        
        next_row = render_token_based_table(
            ws, models, 1, 'thousand', spec_discounts, daily_usages, global_discount
        )
        
        # Should still render headers
        assert ws.cell(row=1, column=1).value == '序号'
        
        # Next row should be 2 (header + no data rows)
        assert next_row == 2
    
    def test_render_token_table_with_discount_and_usage(self):
        """Test complete scenario with discount and daily usage"""
        from app.services.excel_exporter import render_token_based_table
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-max'
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {
            'model1': {
                'spec1': 15.0  # 15% discount
            }
        }
        daily_usages = {
            'model1': {
                'spec1': '2000'  # 2000 tokens per day
            }
        }
        global_discount = 0.0
        
        next_row = render_token_based_table(
            ws, models, 1, 'thousand', spec_discounts, daily_usages, global_discount
        )
        
        # Verify discount is applied
        assert ws.cell(row=2, column=7).value == '15.0%'
        
        # Verify discounted prices
        # 折后输入: 0.04 * 0.85 = 0.034
        assert '¥0.0340/千Token' in str(ws.cell(row=2, column=8).value)
        
        # Verify monthly cost with discount
        # (0.04 + 0.12) * 2000 * 30 * 0.85 = 8160
        monthly_cost = ws.cell(row=2, column=12).value
        assert '¥8,160.00' in monthly_cost
    
    def test_render_token_table_model_id_variations(self):
        """Test handling of different model ID field names"""
        from app.services.excel_exporter import render_token_based_table
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Test with model_code instead of id
        models = [
            {
                'model': {
                    'model_code': 'qwen-max-001',
                    'model_name': 'qwen-max'
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {
            'qwen-max-001': {
                'spec1': 10.0
            }
        }
        daily_usages = {}
        global_discount = 0.0
        
        next_row = render_token_based_table(
            ws, models, 1, 'thousand', spec_discounts, daily_usages, global_discount
        )
        
        # Should successfully apply discount using model_code
        assert ws.cell(row=2, column=7).value == '10.0%'


class TestRenderNonTokenTable:
    """Tests for render_non_token_table() function"""
    
    def test_render_non_token_table_without_discounts(self):
        """Test rendering non-token table without discounts"""
        from openpyxl import Workbook
        from app.services.excel_exporter import render_non_token_table
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data - image generation model
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_code': 'wanx-v1',
                    'model_name': 'Wanx图像生成',
                    'name': 'wanx-v1'
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'model_name': 'wanx-v1',
                        'prices': [
                            {
                                'dimension_code': 'image_count',
                                'unit_price': 0.08
                            }
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {
            'model1': {
                'spec1': '100'
            }
        }
        global_discount = 0.0
        
        # Render table
        next_row = render_non_token_table(
            ws, models, 1, spec_discounts, daily_usages, global_discount
        )
        
        # Verify headers (without discounts)
        expected_headers = ['序号', '模型名称', '单价', '单位', '日估计用量', '预估月用量', '预估月费']
        for col_idx, expected_header in enumerate(expected_headers, 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            assert cell_value == expected_header, f"Header mismatch at column {col_idx}: expected '{expected_header}', got '{cell_value}'"
        
        # Verify data row
        assert ws.cell(row=2, column=1).value == 1  # 序号
        assert ws.cell(row=2, column=2).value == 'Wanx图像生成'  # 模型名称
        assert '¥0.0800' in ws.cell(row=2, column=3).value  # 单价
        assert ws.cell(row=2, column=4).value == '张'  # 单位
        assert ws.cell(row=2, column=5).value == '100'  # 日估计用量
        assert ws.cell(row=2, column=6).value == '3000'  # 预估月用量 (100 * 30)
        assert '¥240.00' in ws.cell(row=2, column=7).value  # 预估月费 (0.08 * 100 * 30)
        
        # Verify next row is correct
        assert next_row == 3
    
    def test_render_non_token_table_with_discounts(self):
        """Test rendering non-token table with discounts"""
        from openpyxl import Workbook
        from app.services.excel_exporter import render_non_token_table
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data - TTS model with discount
        models = [
            {
                'model': {
                    'id': 'model2',
                    'model_code': 'cosyvoice-v1',
                    'model_name': 'CosyVoice语音合成',
                    'name': 'cosyvoice-v1'
                },
                'specs': [
                    {
                        'id': 'spec2',
                        'model_name': 'cosyvoice-v1',
                        'prices': [
                            {
                                'dimension_code': 'character',
                                'unit_price': 0.0001
                            }
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {
            'model2': {
                'spec2': 10.0  # 10% discount
            }
        }
        daily_usages = {
            'model2': {
                'spec2': '50000'
            }
        }
        global_discount = 0.0
        
        # Render table
        next_row = render_non_token_table(
            ws, models, 1, spec_discounts, daily_usages, global_discount
        )
        
        # Verify headers (with discounts)
        expected_headers = ['序号', '模型名称', '单价', '单位', '折扣', '折后单价', '日估计用量', '预估月用量', '预估月费']
        for col_idx, expected_header in enumerate(expected_headers, 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            assert cell_value == expected_header, f"Header mismatch at column {col_idx}: expected '{expected_header}', got '{cell_value}'"
        
        # Verify data row
        assert ws.cell(row=2, column=1).value == 1  # 序号
        assert ws.cell(row=2, column=2).value == 'CosyVoice语音合成'  # 模型名称
        assert '¥0.0001' in ws.cell(row=2, column=3).value  # 单价
        assert ws.cell(row=2, column=4).value == '字符'  # 单位
        assert '10.0%' in ws.cell(row=2, column=5).value  # 折扣
        assert '¥0.0001' in ws.cell(row=2, column=6).value  # 折后单价 (0.0001 * 0.9)
        assert ws.cell(row=2, column=7).value == '50000'  # 日估计用量
        assert ws.cell(row=2, column=8).value == '1500000'  # 预估月用量 (50000 * 30)
        # 预估月费: 0.0001 * 50000 * 30 * 0.9 = 135.00
        assert '¥135.00' in ws.cell(row=2, column=9).value
        
        # Verify next row is correct
        assert next_row == 3
    
    def test_render_non_token_table_multiple_models(self):
        """Test rendering non-token table with multiple models"""
        from openpyxl import Workbook
        from app.services.excel_exporter import render_non_token_table
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data - multiple models
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'Image Model',
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'prices': [
                            {'dimension_code': 'image_count', 'unit_price': 0.08}
                        ]
                    }
                ]
            },
            {
                'model': {
                    'id': 'model2',
                    'model_name': 'Video Model',
                },
                'specs': [
                    {
                        'id': 'spec2',
                        'prices': [
                            {'dimension_code': 'video_second', 'unit_price': 0.5}
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {
            'model1': {'spec1': '100'},
            'model2': {'spec2': '50'}
        }
        global_discount = 0.0
        
        # Render table
        next_row = render_non_token_table(
            ws, models, 1, spec_discounts, daily_usages, global_discount
        )
        
        # Verify we have 2 data rows
        assert ws.cell(row=2, column=1).value == 1  # First model
        assert ws.cell(row=2, column=2).value == 'Image Model'
        assert ws.cell(row=2, column=4).value == '张'  # Unit for image
        
        assert ws.cell(row=3, column=1).value == 2  # Second model
        assert ws.cell(row=3, column=2).value == 'Video Model'
        assert ws.cell(row=3, column=4).value == '秒'  # Unit for video
        
        # Verify next row is correct
        assert next_row == 4
    
    def test_render_non_token_table_missing_price(self):
        """Test rendering non-token table with missing price data"""
        from openpyxl import Workbook
        from app.services.excel_exporter import render_non_token_table
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data - model without price
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'Test Model',
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'prices': []  # No prices
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {}
        global_discount = 0.0
        
        # Render table
        next_row = render_non_token_table(
            ws, models, 1, spec_discounts, daily_usages, global_discount
        )
        
        # Verify data row shows '-' for missing data
        assert ws.cell(row=2, column=3).value == '-'  # 单价
        assert ws.cell(row=2, column=5).value == '-'  # 日估计用量
        assert ws.cell(row=2, column=6).value == '-'  # 预估月用量
        assert ws.cell(row=2, column=7).value == '-'  # 预估月费
        
        # Verify next row is correct
        assert next_row == 3
    
    def test_render_non_token_table_global_discount(self):
        """Test rendering non-token table with global discount"""
        from openpyxl import Workbook
        from app.services.excel_exporter import render_non_token_table
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'ASR Model',
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'prices': [
                            {'dimension_code': 'audio_second', 'unit_price': 0.02}
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {
            'model1': {'spec1': '1000'}
        }
        global_discount = 15.0  # 15% global discount
        
        # Render table
        next_row = render_non_token_table(
            ws, models, 1, spec_discounts, daily_usages, global_discount
        )
        
        # Verify headers include discount columns
        assert ws.cell(row=1, column=5).value == '折扣'
        assert ws.cell(row=1, column=6).value == '折后单价'
        
        # Verify discount is applied
        assert '15.0%' in ws.cell(row=2, column=5).value  # 折扣
        # 折后单价: 0.02 * 0.85 = 0.017
        assert '¥0.0170' in ws.cell(row=2, column=6).value
        # 预估月费: 0.02 * 1000 * 30 * 0.85 = 510.00
        assert '¥510.00' in ws.cell(row=2, column=9).value
        
        # Verify next row is correct
        assert next_row == 3


class TestRenderCategorySection:
    """Tests for render_category_section() function - Requirement 1.3"""
    
    def test_render_category_section_token_based(self):
        """Test rendering a complete token-based category section"""
        from openpyxl import Workbook
        from app.services.excel_exporter import render_category_section
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data for token-based category
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-max',
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ],
                        'remark': '测试备注'
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {}
        global_discount = 0.0
        
        # Render category section
        next_row = render_category_section(
            ws,
            'text_qwen',
            models,
            1,
            'thousand',
            spec_discounts,
            daily_usages,
            global_discount
        )
        
        # Verify category header was rendered (row 1)
        header_cell = ws['A1']
        assert header_cell.value is not None
        assert '💬' in header_cell.value  # Icon
        assert '文本生成-通义千问' in header_cell.value  # Name
        assert '共1项' in header_cell.value  # Item count
        
        # Verify header is merged across columns
        assert 'A1:J1' in ws.merged_cells  # 10 columns for token table without discount
        
        # Verify table headers were rendered (row 2)
        assert ws.cell(row=2, column=1).value == '序号'
        assert ws.cell(row=2, column=2).value == '模型名称'
        assert ws.cell(row=2, column=3).value == '模式'
        assert ws.cell(row=2, column=4).value == 'Token范围'
        assert ws.cell(row=2, column=5).value == '输入单价'
        assert ws.cell(row=2, column=6).value == '输出单价'
        
        # Verify data row was rendered (row 3)
        assert ws.cell(row=3, column=1).value == 1  # 序号
        assert ws.cell(row=3, column=2).value == 'qwen-max'  # 模型名称
        assert ws.cell(row=3, column=3).value == '标准模式'  # 模式
        
        # Verify next row includes spacing (data row + 1 spacing row)
        assert next_row == 5  # Header(1) + Table header(2) + Data(3) + Spacing(4) + Next(5)
    
    def test_render_category_section_non_token_based(self):
        """Test rendering a complete non-token category section"""
        from openpyxl import Workbook
        from app.services.excel_exporter import render_category_section
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data for non-token category (image generation)
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'wanx-v1',
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'prices': [
                            {'dimension_code': 'image_count', 'unit_price': 0.08}
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {}
        global_discount = 0.0
        
        # Render category section
        next_row = render_category_section(
            ws,
            'image_gen',
            models,
            1,
            'thousand',
            spec_discounts,
            daily_usages,
            global_discount
        )
        
        # Verify category header was rendered
        header_cell = ws['A1']
        assert header_cell.value is not None
        assert '🎨' in header_cell.value  # Icon
        assert '图像生成' in header_cell.value  # Name
        assert '共1项' in header_cell.value  # Item count
        
        # Verify header is merged across columns (7 columns for non-token without discount)
        assert 'A1:G1' in ws.merged_cells
        
        # Verify table headers were rendered (row 2)
        assert ws.cell(row=2, column=1).value == '序号'
        assert ws.cell(row=2, column=2).value == '模型名称'
        assert ws.cell(row=2, column=3).value == '单价'
        assert ws.cell(row=2, column=4).value == '单位'
        
        # Verify data row was rendered (row 3)
        assert ws.cell(row=3, column=1).value == 1  # 序号
        assert ws.cell(row=3, column=2).value == 'wanx-v1'  # 模型名称
        assert '¥0.0800' in str(ws.cell(row=3, column=3).value)  # 单价
        assert ws.cell(row=3, column=4).value == '张'  # 单位
        
        # Verify next row includes spacing
        assert next_row == 5
    
    def test_render_category_section_with_discounts(self):
        """Test rendering category section with discounts adds extra columns"""
        from openpyxl import Workbook
        from app.services.excel_exporter import render_category_section
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-plus',
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.02},
                            {'dimension_code': 'output', 'unit_price': 0.06}
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {}
        global_discount = 10.0  # 10% global discount
        
        # Render category section
        next_row = render_category_section(
            ws,
            'text_qwen',
            models,
            1,
            'thousand',
            spec_discounts,
            daily_usages,
            global_discount
        )
        
        # Verify header is merged across 13 columns (token table with discount)
        assert 'A1:M1' in ws.merged_cells
        
        # Verify discount columns are present in table headers
        assert ws.cell(row=2, column=7).value == '折扣'
        assert ws.cell(row=2, column=8).value == '折后输入'
        assert ws.cell(row=2, column=9).value == '折后输出'
        
        # Verify discount is applied in data row
        assert '10.0%' in str(ws.cell(row=3, column=7).value)
    
    def test_render_category_section_multiple_specs(self):
        """Test rendering category section with multiple specs"""
        from openpyxl import Workbook
        from app.services.excel_exporter import render_category_section
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data with 2 models, each with 2 specs
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-max',
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ]
                    },
                    {
                        'id': 'spec2',
                        'mode': '思考模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output_token_thinking', 'unit_price': 0.24}
                        ]
                    }
                ]
            },
            {
                'model': {
                    'id': 'model2',
                    'model_name': 'qwen-plus',
                },
                'specs': [
                    {
                        'id': 'spec3',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.02},
                            {'dimension_code': 'output', 'unit_price': 0.06}
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {}
        global_discount = 0.0
        
        # Render category section
        next_row = render_category_section(
            ws,
            'text_qwen',
            models,
            1,
            'thousand',
            spec_discounts,
            daily_usages,
            global_discount
        )
        
        # Verify category header shows correct item count (3 specs total)
        header_cell = ws['A1']
        assert '共3项' in header_cell.value
        
        # Verify all 3 data rows were rendered
        assert ws.cell(row=3, column=1).value == 1  # First spec
        assert ws.cell(row=3, column=2).value == 'qwen-max'
        
        assert ws.cell(row=4, column=1).value == 2  # Second spec
        assert ws.cell(row=4, column=2).value == 'qwen-max'
        
        assert ws.cell(row=5, column=1).value == 3  # Third spec
        assert ws.cell(row=5, column=2).value == 'qwen-plus'
        
        # Verify next row is correct (header + table header + 3 data rows + spacing)
        assert next_row == 7
    
    def test_render_category_section_with_daily_usage(self):
        """Test rendering category section with daily usage and monthly calculations"""
        from openpyxl import Workbook
        from app.services.excel_exporter import render_category_section
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-max',
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {
            'model1': {'spec1': '1000'}
        }
        global_discount = 0.0
        
        # Render category section
        next_row = render_category_section(
            ws,
            'text_qwen',
            models,
            1,
            'thousand',
            spec_discounts,
            daily_usages,
            global_discount
        )
        
        # Verify daily usage is displayed
        assert ws.cell(row=3, column=7).value == '1000'  # 日估计用量
        
        # Verify monthly usage is calculated (1000 * 30 = 30000)
        assert ws.cell(row=3, column=8).value == '30000'  # 预估月用量
        
        # Verify monthly cost is calculated
        # (0.04 + 0.12) * 1000 * 30 = 4800.00
        monthly_cost = ws.cell(row=3, column=9).value
        assert '¥4,800.00' in str(monthly_cost)
    
    def test_render_category_section_invalid_category_key(self):
        """Test rendering with invalid category key uses defaults"""
        from openpyxl import Workbook
        from app.services.excel_exporter import render_category_section
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'test-model',
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.01},
                            {'dimension_code': 'output', 'unit_price': 0.02}
                        ]
                    }
                ]
            }
        ]
        
        spec_discounts = {}
        daily_usages = {}
        global_discount = 0.0
        
        # Render category section with invalid key
        next_row = render_category_section(
            ws,
            'invalid_category',
            models,
            1,
            'thousand',
            spec_discounts,
            daily_usages,
            global_discount
        )
        
        # Verify header was still rendered with default icon
        header_cell = ws['A1']
        assert header_cell.value is not None
        assert '📋' in header_cell.value  # Default icon
        assert 'invalid_category' in header_cell.value  # Uses key as name
        
        # Verify table was rendered (defaults to token type)
        assert ws.cell(row=2, column=1).value == '序号'
        assert ws.cell(row=3, column=1).value == 1
        
        # Verify function completed successfully
        assert next_row > 1
    
    def test_render_category_section_spacing_row(self):
        """Test that spacing row is added after section"""
        from openpyxl import Workbook
        from app.services.excel_exporter import render_category_section
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Prepare minimal test data
        models = [
            {
                'model': {'id': 'model1', 'model_name': 'test'},
                'specs': [
                    {
                        'id': 'spec1',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.01}
                        ]
                    }
                ]
            }
        ]
        
        # Render category section
        next_row = render_category_section(
            ws, 'text_qwen', models, 1, 'thousand', {}, {}, 0.0
        )
        
        # Verify spacing row is included in next_row calculation
        # Row 1: Header
        # Row 2: Table header
        # Row 3: Data row
        # Row 4: Spacing (empty)
        # Row 5: Next available row
        assert next_row == 5
        
        # Verify row 4 is empty (spacing row)
        assert ws.cell(row=4, column=1).value is None
    
    def test_render_category_section_different_starting_rows(self):
        """Test rendering at different starting row positions"""
        from openpyxl import Workbook
        from app.services.excel_exporter import render_category_section
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data
        models = [
            {
                'model': {'id': 'model1', 'model_name': 'test'},
                'specs': [
                    {
                        'id': 'spec1',
                        'prices': [
                            {'dimension_code': 'image_count', 'unit_price': 0.08}
                        ]
                    }
                ]
            }
        ]
        
        # Render first section starting at row 1
        next_row = render_category_section(
            ws, 'image_gen', models, 1, 'thousand', {}, {}, 0.0
        )
        
        # Verify first section was rendered
        assert ws['A1'].value is not None
        assert next_row == 5
        
        # Render second section starting at returned next_row
        next_row2 = render_category_section(
            ws, 'tts', models, next_row, 'thousand', {}, {}, 0.0
        )
        
        # Verify second section was rendered at correct position
        assert ws[f'A{next_row}'].value is not None
        assert '🔊' in ws[f'A{next_row}'].value  # TTS icon
        assert next_row2 == 9  # 5 + 4 rows for second section
    
    def test_render_category_section_price_unit_million(self):
        """Test rendering with million token price unit"""
        from openpyxl import Workbook
        from app.services.excel_exporter import render_category_section
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Prepare test data
        models = [
            {
                'model': {
                    'id': 'model1',
                    'model_name': 'qwen-max',
                },
                'specs': [
                    {
                        'id': 'spec1',
                        'mode': '标准模式',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ]
                    }
                ]
            }
        ]
        
        # Render with million unit
        next_row = render_category_section(
            ws,
            'text_qwen',
            models,
            1,
            'million',  # Use million instead of thousand
            {},
            {},
            0.0
        )
        
        # Verify prices are converted to million unit
        # 0.04 * 1000 = 40.0
        input_price = ws.cell(row=3, column=5).value
        assert '¥40.0000/百万Token' in str(input_price)
        
        # 0.12 * 1000 = 120.0
        output_price = ws.cell(row=3, column=6).value
        assert '¥120.0000/百万Token' in str(output_price)



class TestGroupModelsByCategory:
    """Tests for group_models_by_category() function - Requirements 7.1, 7.2, 7.3, 8.1, 8.2, 8.3, 8.4"""
    
    def test_group_models_basic(self):
        """Test basic grouping of models by category"""
        from app.services.excel_exporter import group_models_by_category
        
        # Test with a single model
        selected_models = [
            {
                'id': 1,
                'model_code': 'qwen-max',
                'model_name': 'Qwen Max',
                'name': 'Qwen Max'
            }
        ]
        
        model_configs = {
            '1': {
                'variants': [
                    {
                        'id': 1,
                        'model_name': 'qwen-max',
                        'mode': 'thinking',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ],
                        'remark': 'Test remark'
                    }
                ]
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        # Verify structure
        assert 'text_qwen' in result
        assert 'config' in result['text_qwen']
        assert 'models' in result['text_qwen']
        assert len(result['text_qwen']['models']) == 1
        
        # Verify model data
        model_data = result['text_qwen']['models'][0]
        assert 'model' in model_data
        assert 'specs' in model_data
        assert len(model_data['specs']) == 1
        
        # Verify spec data
        spec = model_data['specs'][0]
        assert spec['id'] == 1
        assert spec['model_name'] == 'qwen-max'
        assert spec['mode'] == 'thinking'
        assert spec['token_range'] == '0-128K'
        assert spec['input_price'] == 0.04
        assert spec['output_price'] == 0.12
        assert spec['remark'] == 'Test remark'
    
    def test_group_models_multiple_categories(self):
        """Test grouping models into multiple categories"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {'id': 1, 'model_code': 'qwen-max', 'name': 'Qwen Max'},
            {'id': 2, 'model_code': 'wanx-v1', 'name': 'Wanx V1'},
            {'id': 3, 'model_code': 'cosyvoice', 'name': 'CosyVoice'}
        ]
        
        model_configs = {
            '1': {
                'variants': [
                    {
                        'id': 1,
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04}
                        ]
                    }
                ]
            },
            '2': {
                'variants': [
                    {
                        'id': 2,
                        'prices': [
                            {'dimension_code': 'image_count', 'unit_price': 0.08}
                        ]
                    }
                ]
            },
            '3': {
                'variants': [
                    {
                        'id': 3,
                        'prices': [
                            {'dimension_code': 'character', 'unit_price': 0.001}
                        ]
                    }
                ]
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        # Verify all three categories are present
        assert 'text_qwen' in result
        assert 'image_gen' in result
        assert 'tts' in result
        
        # Verify each category has one model
        assert len(result['text_qwen']['models']) == 1
        assert len(result['image_gen']['models']) == 1
        assert len(result['tts']['models']) == 1
    
    def test_group_models_with_model_code_key(self):
        """Test grouping with model_code as config key (Requirement 8.1)"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {'id': 1, 'model_code': 'qwen-max', 'name': 'Qwen Max'}
        ]
        
        # Use model_code as key instead of id
        model_configs = {
            'qwen-max': {
                'variants': [
                    {
                        'id': 1,
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04}
                        ]
                    }
                ]
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        assert 'text_qwen' in result
        assert len(result['text_qwen']['models']) == 1
    
    def test_group_models_with_legacy_specs_field(self):
        """Test grouping with legacy 'specs' field (Requirement 8.2)"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {'id': 1, 'model_code': 'qwen-max', 'name': 'Qwen Max'}
        ]
        
        # Use legacy 'specs' field instead of 'variants'
        model_configs = {
            '1': {
                'specs': [
                    {
                        'id': 1,
                        'mode': 'standard',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04}
                        ]
                    }
                ]
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        assert 'text_qwen' in result
        assert len(result['text_qwen']['models']) == 1
        assert result['text_qwen']['models'][0]['specs'][0]['mode'] == 'standard'
    
    def test_group_models_with_legacy_spec_field(self):
        """Test grouping with legacy single 'spec' field (Requirement 8.2)"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {'id': 1, 'model_code': 'qwen-max', 'name': 'Qwen Max'}
        ]
        
        # Use legacy single 'spec' field
        model_configs = {
            '1': {
                'spec': {
                    'id': 1,
                    'mode': 'single',
                    'prices': [
                        {'dimension_code': 'input', 'unit_price': 0.04}
                    ]
                }
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        assert 'text_qwen' in result
        assert len(result['text_qwen']['models']) == 1
        assert result['text_qwen']['models'][0]['specs'][0]['mode'] == 'single'
    
    def test_group_models_with_token_range_field(self):
        """Test grouping with legacy 'token_range' field (Requirement 8.3)"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {'id': 1, 'model_code': 'qwen-max', 'name': 'Qwen Max'}
        ]
        
        model_configs = {
            '1': {
                'variants': [
                    {
                        'id': 1,
                        'token_range': '0-32K',  # Legacy field
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04}
                        ]
                    }
                ]
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        assert 'text_qwen' in result
        spec = result['text_qwen']['models'][0]['specs'][0]
        assert spec['token_range'] == '0-32K'
    
    def test_group_models_with_legacy_price_fields(self):
        """Test grouping with legacy input_price/output_price fields (Requirement 8.4)"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {'id': 1, 'model_code': 'qwen-max', 'name': 'Qwen Max'}
        ]
        
        model_configs = {
            '1': {
                'variants': [
                    {
                        'id': 1,
                        'input_price': 0.04,  # Legacy field
                        'output_price': 0.12  # Legacy field
                    }
                ]
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        assert 'text_qwen' in result
        spec = result['text_qwen']['models'][0]['specs'][0]
        assert spec['input_price'] == 0.04
        assert spec['output_price'] == 0.12
    
    def test_group_models_multiple_specs_per_model(self):
        """Test grouping model with multiple specs"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {'id': 1, 'model_code': 'qwen-max', 'name': 'Qwen Max'}
        ]
        
        model_configs = {
            '1': {
                'variants': [
                    {
                        'id': 1,
                        'mode': 'thinking',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04}
                        ]
                    },
                    {
                        'id': 2,
                        'mode': 'standard',
                        'token_tier': '0-32K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.02}
                        ]
                    }
                ]
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        assert 'text_qwen' in result
        assert len(result['text_qwen']['models']) == 1
        assert len(result['text_qwen']['models'][0]['specs']) == 2
        
        # Verify both specs
        specs = result['text_qwen']['models'][0]['specs']
        assert specs[0]['mode'] == 'thinking'
        assert specs[0]['token_range'] == '0-128K'
        assert specs[1]['mode'] == 'standard'
        assert specs[1]['token_range'] == '0-32K'
    
    def test_group_models_with_non_token_prices(self):
        """Test grouping models with non-token pricing"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {'id': 1, 'model_code': 'wanx-v1', 'name': 'Wanx V1'}
        ]
        
        model_configs = {
            '1': {
                'variants': [
                    {
                        'id': 1,
                        'prices': [
                            {'dimension_code': 'image_count', 'unit_price': 0.08}
                        ]
                    }
                ]
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        assert 'image_gen' in result
        spec = result['image_gen']['models'][0]['specs'][0]
        assert spec['non_token_price'] == 0.08
        assert spec['price_unit'] == '张'
        assert spec['dimension_code'] == 'image_count'
        assert spec['input_price'] is None
        assert spec['output_price'] is None
    
    def test_group_models_empty_input(self):
        """Test grouping with empty inputs"""
        from app.services.excel_exporter import group_models_by_category
        
        # Empty selected_models
        result = group_models_by_category([], {})
        assert result == {}
        
        # Empty model_configs
        selected_models = [
            {'id': 1, 'model_code': 'qwen-max', 'name': 'Qwen Max'}
        ]
        result = group_models_by_category(selected_models, {})
        assert result == {}
    
    def test_group_models_missing_config(self):
        """Test grouping when model config is missing"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {'id': 1, 'model_code': 'qwen-max', 'name': 'Qwen Max'},
            {'id': 2, 'model_code': 'wanx-v1', 'name': 'Wanx V1'}
        ]
        
        # Only provide config for model 2
        model_configs = {
            '2': {
                'variants': [
                    {
                        'id': 2,
                        'prices': [
                            {'dimension_code': 'image_count', 'unit_price': 0.08}
                        ]
                    }
                ]
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        # Only model 2 should be grouped
        assert 'image_gen' in result
        assert 'text_qwen' not in result
        assert len(result['image_gen']['models']) == 1
    
    def test_group_models_missing_specs(self):
        """Test grouping when model config has no specs"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {'id': 1, 'model_code': 'qwen-max', 'name': 'Qwen Max'}
        ]
        
        # Config with no specs
        model_configs = {
            '1': {}
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        # Model should be skipped
        assert result == {}
    
    def test_group_models_invalid_spec_data(self):
        """Test grouping with invalid spec data"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {'id': 1, 'model_code': 'qwen-max', 'name': 'Qwen Max'}
        ]
        
        # Config with invalid specs (not a dict)
        model_configs = {
            '1': {
                'variants': [
                    'invalid_spec',  # Not a dict
                    {
                        'id': 2,
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04}
                        ]
                    }
                ]
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        # Only valid spec should be processed
        assert 'text_qwen' in result
        assert len(result['text_qwen']['models']) == 1
        assert len(result['text_qwen']['models'][0]['specs']) == 1
    
    def test_group_models_with_category_field(self):
        """Test grouping respects explicit category field"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {
                'id': 1,
                'model_code': 'custom-model',
                'name': 'Custom Model',
                'category': 'industry'  # Explicit category
            }
        ]
        
        model_configs = {
            '1': {
                'variants': [
                    {
                        'id': 1,
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04}
                        ]
                    }
                ]
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        # Should be in 'industry' category, not 'text_qwen'
        assert 'industry' in result
        assert 'text_qwen' not in result
        assert len(result['industry']['models']) == 1
    
    def test_group_models_normalized_spec_structure(self):
        """Test that normalized spec has all required fields"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {'id': 1, 'model_code': 'qwen-max', 'name': 'Qwen Max'}
        ]
        
        model_configs = {
            '1': {
                'variants': [
                    {
                        'id': 1,
                        'model_name': 'qwen-max',
                        'mode': 'thinking',
                        'token_tier': '0-128K',
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04},
                            {'dimension_code': 'output', 'unit_price': 0.12}
                        ],
                        'remark': 'Test remark'
                    }
                ]
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        spec = result['text_qwen']['models'][0]['specs'][0]
        
        # Verify all required fields are present
        assert 'id' in spec
        assert 'model_name' in spec
        assert 'mode' in spec
        assert 'token_range' in spec
        assert 'input_price' in spec
        assert 'output_price' in spec
        assert 'non_token_price' in spec
        assert 'price_unit' in spec
        assert 'dimension_code' in spec
        assert 'remark' in spec
    
    def test_group_models_default_values(self):
        """Test that missing fields get default values"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {'id': 1, 'model_code': 'qwen-max'}  # Minimal model data
        ]
        
        model_configs = {
            '1': {
                'variants': [
                    {
                        # Minimal spec data - no mode, token_tier, remark
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04}
                        ]
                    }
                ]
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        spec = result['text_qwen']['models'][0]['specs'][0]
        
        # Verify default values
        assert spec['mode'] == '-'
        assert spec['token_range'] == '-'
        assert spec['remark'] == ''
    
    def test_group_models_error_handling(self):
        """Test that errors in processing one model don't affect others"""
        from app.services.excel_exporter import group_models_by_category
        
        selected_models = [
            {'id': 1, 'model_code': 'qwen-max', 'name': 'Qwen Max'},
            {},  # Invalid model - no id or model_code
            {'id': 3, 'model_code': 'wanx-v1', 'name': 'Wanx V1'}
        ]
        
        model_configs = {
            '1': {
                'variants': [
                    {
                        'id': 1,
                        'prices': [
                            {'dimension_code': 'input', 'unit_price': 0.04}
                        ]
                    }
                ]
            },
            '3': {
                'variants': [
                    {
                        'id': 3,
                        'prices': [
                            {'dimension_code': 'image_count', 'unit_price': 0.08}
                        ]
                    }
                ]
            }
        }
        
        result = group_models_by_category(selected_models, model_configs)
        
        # Valid models should still be processed
        assert 'text_qwen' in result
        assert 'image_gen' in result
        assert len(result['text_qwen']['models']) == 1
        assert len(result['image_gen']['models']) == 1
